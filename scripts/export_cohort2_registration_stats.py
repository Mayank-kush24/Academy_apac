"""Export Cohort 2 registration stats to Excel."""
from __future__ import annotations

import os
import re
import sys
from collections import defaultdict
from datetime import date
from pathlib import Path

import pandas as pd
from dotenv import load_dotenv
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from sqlalchemy import create_engine, text

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

from server.routes.dashboard import APAC_COUNTRIES_CANONICAL  # noqa: E402
from server.utils.country_normalize import (  # noqa: E402
    canonical_aliases_lower,
    normalize_country,
)
from server.utils.title_map import _normalize_designation  # noqa: E402

load_dotenv()

AGE_ORDER = ["18-25", "26-35", "36-45", "46-55", "56+"]
OUTPUT = Path(__file__).resolve().parents[1] / "exports" / "cohort_2_registration_stats.xlsx"

# Similar free-text titles roll up to one display label (key = normalized lower).
_DESIGNATION_ALIASES: dict[str, str] = {
    "ceo": "CEO",
    "c.e.o": "CEO",
    "c.e.o.": "CEO",
    "chief executive officer": "CEO",
    "chief executive": "CEO",
    "chief executive officer ceo": "CEO",
    "software engineer": "Software Engineer",
    "software developer": "Software Engineer",
    "software development engineer": "Software Engineer",
    "senior software engineer": "Software Engineer",
    "sr software engineer": "Software Engineer",
    "snr software engineer": "Software Engineer",
    "junior software engineer": "Software Engineer",
    "associate software engineer": "Software Engineer",
    "asst software engineer": "Software Engineer",
    "assistant software engineer": "Software Engineer",
    "software engg": "Software Engineer",
    "software eng": "Software Engineer",
    "sde": "Software Engineer",
    "sde 1": "Software Engineer",
    "sde 2": "Software Engineer",
    "sde 3": "Software Engineer",
    "sde i": "Software Engineer",
    "sde ii": "Software Engineer",
    "sde iii": "Software Engineer",
    "swe": "Software Engineer",
    "software engineer 1": "Software Engineer",
    "software engineer 2": "Software Engineer",
    "software engineer 3": "Software Engineer",
    "software engineer 4": "Software Engineer",
    "software developer 1": "Software Engineer",
    "software developer 2": "Software Engineer",
    "software developer 3": "Software Engineer",
    "founder": "Founder / Co-Founder",
    "co founder": "Founder / Co-Founder",
    "cofounder": "Founder / Co-Founder",
    "co-founder": "Founder / Co-Founder",
    "founder and ceo": "Founder / Co-Founder",
    "founder ceo": "Founder / Co-Founder",
    "co founder and ceo": "Founder / Co-Founder",
    "managing director": "Managing Director",
    "md": "Managing Director",
    "m.d": "Managing Director",
    "m.d.": "Managing Director",
    "director": "Director",
    "product manager": "Product Manager",
    "product owner": "Product Owner",
    "project manager": "Project Manager",
    "account manager": "Account Manager",
    "business analyst": "Business Analyst",
    "business development manager": "Business Development Manager",
    "owner": "Owner",
    "data engineer": "Data Engineer",
    "technical consultant": "Technical Consultant",
    "manager": "Manager",
    "developer": "Software Engineer",
}


def _title_key(s: str) -> str:
    return " ".join(s.strip().lower().split())


def _pick_display(counts: dict[str, int]) -> str:
    return sorted(counts.items(), key=lambda x: (-x[1], x[0]))[0][0]


def canonicalize_designation(raw: str | None) -> str | None:
    """Merge near-duplicate titles (CEO/Chief Executive Officer, SE(1)/SE(2), etc.)."""
    if raw is None:
        return None
    if _title_key(raw) == "student":
        return "Student"

    cleaned = _normalize_designation(raw)  # strips e.g. "( 2 )"
    if not cleaned:
        return None

    key = _title_key(cleaned)
    key = re.sub(r"[./]", "", key)
    key = re.sub(r"\s+", " ", key).strip()
    # Drop trailing bare level markers left after paren strip: "software engineer 2"
    key = re.sub(r"\s+[ivx\d]+$", "", key).strip()

    if key in _DESIGNATION_ALIASES:
        return _DESIGNATION_ALIASES[key]

    # Soft match: prefer longer alias phrases (e.g. "software engineer" before "developer")
    for alias_key, label in sorted(
        _DESIGNATION_ALIASES.items(), key=lambda x: len(x[0]), reverse=True
    ):
        if key.startswith(alias_key + " "):
            return label

    return cleaned.title() if cleaned.isupper() or cleaned.islower() else cleaned


def _autosize(ws) -> None:
    for col in ws.columns:
        letter = get_column_letter(col[0].column)
        width = max((len(str(c.value)) if c.value is not None else 0) for c in col)
        ws.column_dimensions[letter].width = min(max(width + 2, 10), 70)
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")


def main() -> int:
    database_url = os.getenv("DATABASE_URL")
    if not database_url:
        print("[ERROR] DATABASE_URL is not set")
        return 1

    engine = create_engine(database_url)
    today = date.today().isoformat()
    india_aliases = list(canonical_aliases_lower(["India"]))
    apac_aliases = list(canonical_aliases_lower(APAC_COUNTRIES_CANONICAL))

    with engine.connect() as conn:
        total = int(
            conn.execute(text("SELECT COUNT(*) FROM cohort_2_user_pii_combined")).scalar()
        )
        india_total = int(
            conn.execute(
                text(
                    """
                    SELECT COUNT(*) FROM cohort_2_user_pii_combined
                    WHERE LOWER(TRIM(country)) = ANY(:india)
                    """
                ),
                {"india": india_aliases},
            ).scalar()
        )
        apac_excl = int(
            conn.execute(
                text(
                    """
                    SELECT COUNT(*) FROM cohort_2_user_pii_combined
                    WHERE LOWER(TRIM(country)) = ANY(:apac)
                      AND LOWER(TRIM(country)) <> ALL(:india)
                    """
                ),
                {"apac": apac_aliases, "india": india_aliases},
            ).scalar()
        )
        student_count = int(
            conn.execute(
                text(
                    """
                    SELECT COUNT(*) FROM cohort_2_user_pii_combined
                    WHERE occupation IS NOT NULL AND LOWER(occupation) LIKE '%student%'
                    """
                )
            ).scalar()
        )

        india_rows = conn.execute(
            text(
                """
                SELECT TRIM(city) AS city, COUNT(*) AS n
                FROM cohort_2_user_pii_combined
                WHERE city IS NOT NULL AND TRIM(city) <> ''
                  AND LOWER(TRIM(country)) = ANY(:india)
                GROUP BY TRIM(city)
                ORDER BY n DESC
                LIMIT 50
                """
            ),
            {"india": india_aliases},
        ).fetchall()

        apac_rows = conn.execute(
            text(
                """
                SELECT TRIM(city) AS city, TRIM(country) AS country, COUNT(*) AS n
                FROM cohort_2_user_pii_combined
                WHERE city IS NOT NULL AND TRIM(city) <> ''
                  AND country IS NOT NULL AND TRIM(country) <> ''
                  AND LOWER(TRIM(country)) = ANY(:apac)
                  AND LOWER(TRIM(country)) <> ALL(:india)
                GROUP BY TRIM(city), TRIM(country)
                ORDER BY n DESC
                LIMIT 200
                """
            ),
            {"apac": apac_aliases, "india": india_aliases},
        ).fetchall()

        age_rows = conn.execute(
            text(
                f"""
                WITH ages AS (
                  SELECT EXTRACT(YEAR FROM AGE(DATE '{today}', date_of_birth))::int AS age
                  FROM cohort_2_user_pii_combined
                  WHERE date_of_birth IS NOT NULL
                )
                SELECT
                  CASE
                    WHEN age BETWEEN 18 AND 25 THEN '18-25'
                    WHEN age BETWEEN 26 AND 35 THEN '26-35'
                    WHEN age BETWEEN 36 AND 45 THEN '36-45'
                    WHEN age BETWEEN 46 AND 55 THEN '46-55'
                    WHEN age > 55 THEN '56+'
                    ELSE NULL
                  END AS age_group,
                  COUNT(*) AS n
                FROM ages
                WHERE age >= 18
                GROUP BY 1
                """
            )
        ).fetchall()

        desig_rows = conn.execute(
            text(
                """
                SELECT TRIM(designation) AS desig, COUNT(*) AS n
                FROM cohort_2_user_pii_combined
                WHERE designation IS NOT NULL AND TRIM(designation) <> ''
                  AND (
                    occupation IS NULL
                    OR LOWER(occupation) NOT LIKE '%student%'
                  )
                GROUP BY 1
                """
            )
        ).fetchall()

        college_rows = conn.execute(
            text(
                """
                SELECT TRIM(organization_name) AS org, COUNT(*) AS n
                FROM cohort_2_user_pii_combined
                WHERE organization_name IS NOT NULL AND TRIM(organization_name) <> ''
                  AND occupation IS NOT NULL AND LOWER(occupation) LIKE '%student%'
                GROUP BY 1
                ORDER BY n DESC
                LIMIT 800
                """
            )
        ).fetchall()

        company_rows = conn.execute(
            text(
                """
                SELECT TRIM(organization_name) AS org, COUNT(*) AS n
                FROM cohort_2_user_pii_combined
                WHERE organization_name IS NOT NULL AND TRIM(organization_name) <> ''
                  AND (
                    occupation IS NULL
                    OR LOWER(occupation) NOT LIKE '%student%'
                  )
                GROUP BY 1
                ORDER BY n DESC
                LIMIT 800
                """
            )
        ).fetchall()

    # --- India cities ---
    india_merged: dict[str, dict[str, int]] = defaultdict(lambda: defaultdict(int))
    india_totals: dict[str, int] = defaultdict(int)
    for city, n in india_rows:
        k = _title_key(city)
        india_merged[k][city] += int(n)
        india_totals[k] += int(n)
    india_cities = [
        {"Rank": i + 1, "City": _pick_display(india_merged[k]), "Registrations": india_totals[k]}
        for i, k in enumerate(sorted(india_totals, key=lambda x: -india_totals[x])[:20])
    ]

    # --- APAC cities ---
    apac_merged: dict[tuple[str, str], dict[str, int]] = defaultdict(lambda: defaultdict(int))
    apac_totals: dict[tuple[str, str], int] = defaultdict(int)
    for city, country, n in apac_rows:
        cty = normalize_country(country) or country
        k = (_title_key(city), cty)
        apac_merged[k][city] += int(n)
        apac_totals[k] += int(n)
    apac_cities = [
        {
            "Rank": i + 1,
            "City": _pick_display(apac_merged[k]),
            "Country": k[1],
            "Registrations": apac_totals[k],
        }
        for i, k in enumerate(sorted(apac_totals, key=lambda x: -apac_totals[x])[:20])
    ]

    # --- Age ---
    age_map = {g: int(n) for g, n in age_rows if g}
    age_total = sum(age_map.values())
    age = [
        {
            "Age Group": g,
            "Registrations": age_map.get(g, 0),
            "% of Known 18+": round(100.0 * age_map.get(g, 0) / age_total, 1) if age_total else 0,
        }
        for g in AGE_ORDER
        if age_map.get(g, 0) > 0
    ]

    # --- Designations (similar titles rolled up; Student excluded) ---
    desig_totals: dict[str, int] = defaultdict(int)
    for desig, n in desig_rows:
        label = canonicalize_designation(desig)
        if not label or _title_key(label) == "student":
            continue
        desig_totals[label] += int(n)
    designations = [
        {
            "Rank": i + 1,
            "Designation": label,
            "Registrations": desig_totals[label],
        }
        for i, label in enumerate(
            sorted(desig_totals, key=lambda x: (-desig_totals[x], x.lower()))[:25]
        )
    ]

    def _top_orgs(rows, label_col: str, limit: int = 100):
        merged: dict[str, dict[str, int]] = defaultdict(lambda: defaultdict(int))
        totals: dict[str, int] = defaultdict(int)
        skip = {
            "individual",
            "na",
            "n/a",
            "none",
            "nil",
            "self",
            "self employed",
            "self-employed",
            "freelancer",
            "freelance",
            "not applicable",
            "student",
            "-",
            "--",
            ".",
        }
        for org, n in rows:
            if not org:
                continue
            k = _title_key(org)
            if not k or k in skip:
                continue
            merged[k][org] += int(n)
            totals[k] += int(n)
        return [
            {
                "Rank": i + 1,
                label_col: _pick_display(merged[k]),
                "Registrations": totals[k],
            }
            for i, k in enumerate(sorted(totals, key=lambda x: -totals[x])[:limit])
        ]

    colleges = _top_orgs(college_rows, "College")
    companies = _top_orgs(company_rows, "Company")

    summary = [
        {"Metric": "As of", "Value": today},
        {"Metric": "Source", "Value": "cohort_2_user_pii_combined"},
        {"Metric": "Total registrations", "Value": total},
        {"Metric": "India registrations", "Value": india_total},
        {"Metric": "APAC excl. India registrations", "Value": apac_excl},
        {"Metric": "Student registrations", "Value": student_count},
        {"Metric": "Age known (18+)", "Value": age_total},
        {
            "Metric": "Notes",
            "Value": (
                "APAC cities exclude India; country aliases rolled up; "
                "Student excluded from designations; "
                "similar designations merged (e.g. CEO/Chief Executive Officer, "
                "Software Engineer variants, Founder/Co-Founder); "
                "Top 100 Colleges = student organization_name; "
                "Top 100 Companies = non-student organization_name"
            ),
        },
    ]

    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(OUTPUT, engine="openpyxl") as writer:
        pd.DataFrame(summary).to_excel(writer, sheet_name="Summary", index=False)
        pd.DataFrame(india_cities).to_excel(
            writer, sheet_name="Top 20 Cities - India", index=False
        )
        pd.DataFrame(apac_cities).to_excel(
            writer, sheet_name="Top 20 Cities - APAC", index=False
        )
        pd.DataFrame(age).to_excel(writer, sheet_name="Age Composition", index=False)
        pd.DataFrame(designations).to_excel(
            writer, sheet_name="Top Designations", index=False
        )
        pd.DataFrame(companies).to_excel(
            writer, sheet_name="Top 100 Companies", index=False
        )
        pd.DataFrame(colleges).to_excel(
            writer, sheet_name="Top 100 Colleges", index=False
        )
        for name in writer.book.sheetnames:
            _autosize(writer.book[name])

    print(f"Wrote {OUTPUT}")
    print("Top designations:", [d["Designation"] for d in designations[:5]])
    print("Top companies:", [c["Company"] for c in companies[:5]])
    print("Top colleges:", [c["College"] for c in colleges[:5]])
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
