"""
Excel parsing and field mapping utilities
"""
from collections import Counter
from types import SimpleNamespace
import re

import pandas as pd
from datetime import datetime
from sqlalchemy.exc import DataError, IntegrityError


def _import_models():
    """Participant ORM classes for the current cohort (g.table_prefix)."""
    from server import models as m
    from server.utils.cohort_participant_models import participant_model

    return SimpleNamespace(
        UserPII=participant_model(m.UserPII),
        UserPIIInjected=participant_model(m.UserPIIInjected),
        SkillboostProfile=participant_model(m.SkillboostProfile),
        SkillLabSubmission=participant_model(m.SkillLabSubmission),
        CodeLabSubmission=participant_model(m.CodeLabSubmission),
        ProjectSubmission=participant_model(m.ProjectSubmission),
        OptionalMcqResponse=participant_model(m.OptionalMcqResponse),
        MainMcqResponse=participant_model(m.MainMcqResponse),
    )


# ─────────────────────────────────────────────────────────────────────────────
# Sheet-detection registry (single source of truth)
#
# Detection is regex-based and prefix-agnostic so Action Center can renumber
# tabs (e.g. "16.MCQ Track 1" -> "15.MCQ Track 1") without breaking imports.
# Each detector classifies any workbook tab whose name contains the pattern.
# Main MCQ uses a negative lookahead so "MCQ Optional Track 1" is NEVER
# misclassified as Main MCQ.
# ─────────────────────────────────────────────────────────────────────────────


class SheetDetector:
    """One sheet-classification rule.

    module           — short stable key the rest of the code groups by
                       (e.g. "main_mcq", "optional_mcq", "skillboost_profile")
    label            — human-readable name shown in the UI
    patterns         — list of regex strings (case-insensitive); first hit wins
    track / lab      — optional track/lab number when the module is per-track
    critical         — bool: when True, absence of this sheet may be surfaced as a
                       missing-critical warning (see critical_cohorts).
    cohorts          — optional iterable of cohort_ids this detector applies to
                       (None = applies to all cohorts).
    critical_cohorts — optional iterable limiting which cohorts treat this sheet as
                       critical when missing (None = same as applies_to_cohort).
    """

    __slots__ = ("module", "label", "patterns", "track", "lab",
                 "critical", "cohorts", "critical_cohorts", "_compiled")

    def __init__(self, module, label, patterns, track=None, lab=None,
                 critical=False, cohorts=None, critical_cohorts=None):
        self.module = module
        self.label = label
        self.patterns = list(patterns)
        self.track = track
        self.lab = lab
        self.critical = bool(critical)
        self.cohorts = tuple(cohorts) if cohorts is not None else None
        self.critical_cohorts = (
            tuple(critical_cohorts) if critical_cohorts is not None else None
        )
        self._compiled = [re.compile(p, re.IGNORECASE) for p in self.patterns]

    @staticmethod
    def _effective_cohort_id(cohort_id):
        """Legacy routes use cohort_id=None; treat that as Cohort 1."""
        return 1 if cohort_id is None else cohort_id

    def applies_to_cohort(self, cohort_id):
        """True if this detector is active for the given cohort_id (None = any)."""
        if self.cohorts is None:
            return True
        return self._effective_cohort_id(cohort_id) in self.cohorts

    def matches(self, sheet_name):
        """True if sheet_name (whitespace-collapsed) matches any compiled pattern."""
        if sheet_name is None:
            return False
        norm = re.sub(r"\s+", " ", str(sheet_name).replace("\u00a0", " ")).strip()
        return any(rx.search(norm) for rx in self._compiled)

    def is_critical_for_cohort(self, cohort_id):
        if not self.critical:
            return False
        effective = self._effective_cohort_id(cohort_id)
        if self.critical_cohorts is not None:
            return effective in self.critical_cohorts
        if self.cohorts is not None:
            return effective in self.cohorts
        return True


# Order matters: more specific detectors first so they "win" over fall-throughs.
# Main MCQ uses negative lookahead "(?!optional\s)" so a tab named
# "MCQ Optional Track 1" is matched by optional_mcq, never main_mcq.
SHEET_DETECTORS = [
    # Skills Boost profile (any of three accepted phrasings)
    SheetDetector(
        "skillboost_profile", "Skills Boost Profile",
        [r"share\s+your\s+google\s+skills\s+(?:pub|boost|pu)\b"],
        critical=False,
    ),

    # Skill Lab Submission (detect in any cohort; required only for Cohort 1 exports)
    SheetDetector(
        "skilllab_submission", "Skill Lab Submission",
        [r"google\s+skills\s+lab\s+submissio"],
        critical=True,
        critical_cohorts=(1,),
    ),

    # Code Lab Submission (Cohort 1 Action Center export; pattern does not match Cohort 2 tab names)
    SheetDetector(
        "codelab_submission", "Code Lab Submission",
        [r"code\s+lab\s+submissio"],
        critical=True,
        cohorts=(1,),
    ),

    # Cohort 2 Code Lab — Student / Professional track tabs (prefix-agnostic)
    # Student Track: track_number is assigned as 3 at import time (profiles.js convention).
    # Detector track stays None so it does not collide with Professional Track 3 in SHEET_DETECTORS.
    SheetDetector(
        "codelab_submission", "Student Track Codelab Building",
        [r"student\s*track.*codelab\s*buildi"],
        critical=True,
        cohorts=(2,),
    ),
    SheetDetector(
        "codelab_submission", "Professional Track 1 Codelab",
        [r"professional\s+track\s+1\s+codelab"],
        track=1,
        critical=True,
        cohorts=(2,),
    ),
    SheetDetector(
        "codelab_submission", "Professional Track 2 Codelab",
        [r"professional\s+track\s+2\s+codelab"],
        track=2,
        critical=True,
        cohorts=(2,),
    ),
    SheetDetector(
        "codelab_submission", "Professional Track 3 Codelab",
        [r"professional\s+track\s+3\s+codelab"],
        track=3,
        critical=True,
        cohorts=(2,),
    ),

    # Optional MCQ Track N (must contain "Optional" between MCQ and Track)
    SheetDetector("optional_mcq", "Optional MCQ Track 1",
                  [r"mcq\s+optional\s+track\s+1\b"], track=1, critical=True, cohorts=(1,)),
    SheetDetector("optional_mcq", "Optional MCQ Track 2",
                  [r"mcq\s+optional\s+track\s+2\b"], track=2, critical=True, cohorts=(1,)),
    SheetDetector("optional_mcq", "Optional MCQ Track 3",
                  [r"mcq\s+optional\s+track\s+3\b"], track=3, critical=True, cohorts=(1,)),

    # Main MCQ Track N (must NOT have "Optional" between MCQ and Track).
    # Negative lookahead protects against "MCQ Optional Track 1" matching here.
    SheetDetector("main_mcq", "Main MCQ Track 1",
                  [r"mcq\s+(?!optional\s)track\s+1\b"], track=1, critical=True, cohorts=(1,)),
    SheetDetector("main_mcq", "Main MCQ Track 2",
                  [r"mcq\s+(?!optional\s)track\s+2\b"], track=2, critical=True, cohorts=(1,)),
    SheetDetector("main_mcq", "Main MCQ Track 3",
                  [r"mcq\s+(?!optional\s)track\s+3\b"], track=3, critical=True, cohorts=(1,)),

    # Cohort 2/3 Main MCQ — Professional / Student track quiz tabs (separate Quiz workbook).
    # Not critical: Forms exports often omit these; Quiz files are imported separately.
    SheetDetector(
        "main_mcq", "Professional Track 1 MCQ",
        [r"professional\s+track\s+1\s+mcq"],
        track=1, critical=False, cohorts=(2, 3),
    ),
    SheetDetector(
        "main_mcq", "Professional Track 2 MCQ",
        [r"professional\s+track\s+2\s+mcq"],
        track=2, critical=False, cohorts=(2, 3),
    ),
    SheetDetector(
        "main_mcq", "Professional Track 3 MCQ",
        [r"professional\s+track\s+3\s+mcq"],
        track=3, critical=False, cohorts=(2, 3),
    ),
    # Student Track → track_number=3 (same convention as Code Lab / profiles grid).
    SheetDetector(
        "main_mcq", "Student Track MCQ",
        [r"student\s+track\s+mcq"],
        track=3, critical=False, cohorts=(2, 3),
    ),

    # Lab Completion N x Track N
    SheetDetector("lab_completion", "Lab Completion 1 Track 1",
                  [r"lab\s+completion\s+1\s+track\s+1\b"], lab=1, track=1, critical=True, cohorts=(1,)),
    SheetDetector("lab_completion", "Lab Completion 1 Track 2",
                  [r"lab\s+completion\s+1\s+track\s+2\b"], lab=1, track=2, critical=True, cohorts=(1,)),
    SheetDetector("lab_completion", "Lab Completion 1 Track 3",
                  [r"lab\s+completion\s+1\s+track\s+3\b"], lab=1, track=3, critical=True, cohorts=(1,)),
    SheetDetector("lab_completion", "Lab Completion 2 Track 1",
                  [r"lab\s+completion\s+2\s+track\s+1\b"], lab=2, track=1, critical=True, cohorts=(1,)),
    SheetDetector("lab_completion", "Lab Completion 2 Track 2",
                  [r"lab\s+completion\s+2\s+track\s+2\b"], lab=2, track=2, critical=True, cohorts=(1,)),
    SheetDetector("lab_completion", "Lab Completion 2 Track 3",
                  [r"lab\s+completion\s+2\s+track\s+3\b"], lab=2, track=3, critical=True, cohorts=(1,)),

    # Project Submission Track N
    SheetDetector("project_submission", "Project Submission Track 1",
                  [r"project\s+submission\s+track\s+1\b"], track=1, critical=True, cohorts=(1,)),
    SheetDetector("project_submission", "Project Submission Track 2",
                  [r"project\s+submission\s+track\s+2\b"], track=2, critical=True, cohorts=(1,)),
    SheetDetector("project_submission", "Project Submission Track 3",
                  [r"project\s+submission\s+track\s+3\b"], track=3, critical=True, cohorts=(1,)),

    # Cohort 2 Optional MCQ — anchored exact pattern "<digits>.MCQ Optional"
    # (e.g. "14.MCQ Optional"). Active only for Cohort 2.
    SheetDetector(
        "cohort2_optional_mcq", "Cohort 2 Optional MCQ",
        [r"^\s*\d+\.\s*mcq\s+optional\s*$"],
        critical=True, cohorts=(2,),
    ),
]


# ─────────────────────────────────────────────────────────────────────────────
# Backward-compatibility constants
# Existing call sites still import these names; they now resolve to the
# regex-based detectors above. A new substring such as "MCQ Track 1" works
# regardless of any leading "16." / "15." numbering.
# ─────────────────────────────────────────────────────────────────────────────
SKILLBOOST_SHEET_SUBSTRINGS = (
    "Share your Google Skills Pub",
    "Share your Google Skills Boost",
    "Share your Google Skills Pu",
)
SKILLBOOST_SHEET_SUBSTRING = SKILLBOOST_SHEET_SUBSTRINGS[0]
SKILLLAB_SUBMISSION_SHEET_SUBSTRING = "Google Skills Lab Submissio"
CODELAB_SUBMISSION_SHEET_SUBSTRING = "Code Lab Submissio"


def _codelab_sheet_track_number(sheet_name, track):
    """
    Resolve track_number for a Code Lab sheet.
    Cohort 2 Student Track → 3 (profiles / verification UI convention).
    """
    if track is not None:
        return track
    if "student" in (sheet_name or "").lower():
        return 3
    return None


def _codelab_default_problem_statement(sheet_name, track):
    """
    Cohort 2 Code Lab tabs have no Problem Statement column — lab identity is the tab.
    Returns a stable label for problem_statement (and Lab column in the UI).
    """
    name = (sheet_name or "").lower()
    if "student" in name:
        return "Student Track Codelab"
    if track == 1 or ("professional" in name and "track 1" in name):
        return "Professional Track 1 Codelab"
    if track == 2 or ("professional" in name and "track 2" in name):
        return "Professional Track 2 Codelab"
    if track == 3 or ("professional" in name and "track 3" in name):
        return "Professional Track 3 Codelab"
    return None
LAB_COMPLETION_SHEET_SUBSTRINGS = [
    {'substring': 'Lab Completion 1 Track 1', 'lab': 1, 'track': 1},
    {'substring': 'Lab Completion 1 Track 2', 'lab': 1, 'track': 2},
    {'substring': 'Lab Completion 1 Track 3', 'lab': 1, 'track': 3},
    {'substring': 'Lab Completion 2 Track 1', 'lab': 2, 'track': 1},
    {'substring': 'Lab Completion 2 Track 2', 'lab': 2, 'track': 2},
    {'substring': 'Lab Completion 2 Track 3', 'lab': 2, 'track': 3},
]
MCQ_OPTIONAL_TRACK1_SHEET_SUBSTRING = "MCQ Optional Track 1"
MCQ_OPTIONAL_TRACK2_SHEET_SUBSTRING = "MCQ Optional Track 2"
MCQ_OPTIONAL_TRACK3_SHEET_SUBSTRING = "MCQ Optional Track 3"
_COHORT2_OPTIONAL_MCQ_SHEET_NAME_RE = re.compile(
    r"^\s*\d+\.\s*MCQ\s+Optional\s*$",
    re.IGNORECASE,
)
# Prefix-agnostic Main MCQ substrings (no leading "16."/"17."/"18." required).
MCQ_MAIN_TRACK1_SHEET_SUBSTRING = "MCQ Track 1"
MCQ_MAIN_TRACK2_SHEET_SUBSTRING = "MCQ Track 2"
MCQ_MAIN_TRACK3_SHEET_SUBSTRING = "MCQ Track 3"
PROJECT_SUBMISSION_SHEET_DEFINITIONS = [
    {'track': 1, 'substring': 'Project Submission Track 1'},
    {'track': 2, 'substring': 'Project Submission Track 2'},
    {'track': 3, 'substring': 'Project Submission Track 3'},
]

# Per-row SAVEPOINT (db.session.begin_nested) is used inside every importer so
# a single failed row never aborts a multi-row batch. The outer commit at the
# end of the importer flushes everything that survived.
IMPORT_BATCH_COMMIT_SIZE = 500
# Chunk size for SQL IN (...) when preloading rows by email
IMPORT_EMAIL_IN_CHUNK = 500


def _match_detector(sheet_name, cohort_id=None):
    """Return the first SheetDetector matching sheet_name (active for cohort), or None."""
    if sheet_name is None:
        return None
    for det in SHEET_DETECTORS:
        if not det.applies_to_cohort(cohort_id):
            continue
        if det.matches(sheet_name):
            return det
    return None


def classify_workbook_sheets(file_path, cohort_id=None):
    """Walk every tab in the workbook and classify it via SHEET_DETECTORS.

    Returns dict with:
        sheet_count
        sheet_names         — every tab in workbook order
        sheet_mapping       — list of dicts, one per tab in workbook order:
                              {"sheet_name", "module", "label", "track",
                               "lab", "status"}
                              status in {"detected", "duplicate", "unrecognised"}.
                              "duplicate" = a later tab matched a detector that
                              an earlier tab already claimed (only the first
                              wins for import).
        detected_by_module  — dict module_key -> list of mapping rows that
                              were classified as that module (in workbook order).
        missing_critical_modules — list of {"module", "label", "track", "lab"}
                              for every critical detector (for the cohort) that
                              did not match any tab.
    """
    sheet_names = get_sheet_names(file_path)
    sheet_mapping = []
    detected_by_module = {}
    matched_detector_keys = set()

    for name in sheet_names:
        det = _match_detector(name, cohort_id=cohort_id)
        if det is None:
            sheet_mapping.append({
                "sheet_name": name,
                "module": None,
                "label": None,
                "track": None,
                "lab": None,
                "status": "unrecognised",
            })
            continue

        det_key = (det.module, det.track, det.lab)
        status = "duplicate" if det_key in matched_detector_keys else "detected"
        row = {
            "sheet_name": name,
            "module": det.module,
            "label": det.label,
            "track": det.track,
            "lab": det.lab,
            "status": status,
        }
        sheet_mapping.append(row)
        if status == "detected":
            matched_detector_keys.add(det_key)
            detected_by_module.setdefault(det.module, []).append(row)

    # Missing critical modules
    missing = []
    for det in SHEET_DETECTORS:
        if not det.is_critical_for_cohort(cohort_id):
            continue
        det_key = (det.module, det.track, det.lab)
        if det_key in matched_detector_keys:
            continue
        missing.append({
            "module": det.module,
            "label": det.label,
            "track": det.track,
            "lab": det.lab,
        })

    return {
        "sheet_count": len(sheet_names),
        "sheet_names": sheet_names,
        "sheet_mapping": sheet_mapping,
        "detected_by_module": detected_by_module,
        "missing_critical_modules": missing,
    }


def first_sheet_for_module(classification, module, track=None, lab=None):
    """Return the first detected sheet name for (module, track, lab), or None."""
    for row in classification.get("detected_by_module", {}).get(module, []) or []:
        if track is not None and row.get("track") != track:
            continue
        if lab is not None and row.get("lab") != lab:
            continue
        return row.get("sheet_name")
    return None


def _chunk_list(items, size):
    """Yield lists of at most `size` elements from iterable."""
    buf = list(items)
    for i in range(0, len(buf), size):
        yield buf[i:i + size]


def get_sheet_names(file_path):
    """Return list of worksheet names and count for the given Excel file."""
    try:
        if str(file_path).lower().endswith('.xlsx'):
            import openpyxl
            wb = openpyxl.load_workbook(file_path, read_only=True)
            names = list(wb.sheetnames)
            wb.close()
            return names
        xl = pd.ExcelFile(file_path)
        return list(xl.sheet_names)
    except Exception:
        return []


def find_skillboost_sheet_name_in_list(sheet_names):
    """
    First worksheet whose name contains any of SKILLBOOST_SHEET_SUBSTRINGS.
    Matches in substring order, then workbook sheet order (same as previous single-substring behavior).
    """
    if not sheet_names:
        return None
    for sub in SKILLBOOST_SHEET_SUBSTRINGS:
        for name in sheet_names:
            if name is not None and sub.lower() in str(name).lower():
                return name
    return None


def find_skillboost_sheet_in_file(file_path):
    return find_skillboost_sheet_name_in_list(get_sheet_names(file_path))


def skillboost_sheet_not_found_message():
    """Short user-facing hint when no Skills profile sheet is detected."""
    inner = '", "'.join(SKILLBOOST_SHEET_SUBSTRINGS)
    return f'No worksheet whose name contains one of: "{inner}".'


ACTION_CENTER_NO_RECOGNIZED_SHEETS_MESSAGE = (
    "No recognized Action Center subsheets in this workbook. "
    "Expected at least one sheet whose name contains one of (any leading "
    'numbering is ignored): "MCQ Track 1/2/3" (Main MCQ), '
    '"MCQ Optional Track 1/2/3" (Optional MCQ), '
    '"Share your Google Skills …" (profile), '
    '"Google Skills Lab Submissio…", "Code Lab Submissio…" (Cohort 1), '
    '"Student Track Codelab Buildi…" / "Professional Track N Codelab" (Cohort 2), '
    '"Lab Completion 1/2 Track 1/2/3", or "Project Submission Track 1/2/3".'
)


def _sheets_has_codelab_data(sheets):
    """True if sheets dict has at least one non-empty Code Lab worksheet loaded."""
    if not sheets:
        return False
    if sheets.get("codelab_sheet_name") and _nonempty_dataframe(sheets.get("codelab_df")):
        return True
    for cl in sheets.get("codelab_submission_sheets") or []:
        if _nonempty_dataframe(cl.get("df")):
            return True
    return False


def _nonempty_dataframe(df):
    return df is not None and len(df) > 0


def action_center_has_non_profile_imports(sheets):
    """
    True if load_all_skillboost_sheets found at least one non-empty importable sheet
    other than the Skill Lab profile sheet (submission, code lab, MCQ, labs, projects).
    """
    if not sheets or not isinstance(sheets, dict):
        return False

    if sheets.get("submission_sheet_name") and _nonempty_dataframe(sheets.get("submission_df")):
        return True
    if _sheets_has_codelab_data(sheets):
        return True
    for lc in sheets.get("lab_completion_sheets") or []:
        if _nonempty_dataframe(lc.get("df")):
            return True
    for m in sheets.get("main_mcq_sheets") or []:
        if _nonempty_dataframe(m.get("df")):
            return True
    for om in sheets.get("optional_mcq_sheets") or []:
        if _nonempty_dataframe(om.get("df")):
            return True
    for ps in sheets.get("project_submission_sheets") or []:
        if _nonempty_dataframe(ps.get("df")):
            return True
    c2 = sheets.get("cohort2_optional_mcq_sheet")
    if c2 and isinstance(c2, dict) and _nonempty_dataframe(c2.get("df")):
        return True
    return False


def cohort2_action_center_optional_mcq_only(sheets):
    """
    True when the workbook has Cohort 2 Optional MCQ data and no other importable
    non-profile subsheets (e.g. Code Lab Submission, Skill Lab Submission, Main MCQ).
    """
    c2 = sheets.get("cohort2_optional_mcq_sheet") if sheets else None
    if not (c2 and isinstance(c2, dict) and _nonempty_dataframe(c2.get("df"))):
        return False

    if sheets.get("submission_sheet_name") and _nonempty_dataframe(sheets.get("submission_df")):
        return False
    if _sheets_has_codelab_data(sheets):
        return False
    for lc in sheets.get("lab_completion_sheets") or []:
        if _nonempty_dataframe(lc.get("df")):
            return False
    for m in sheets.get("main_mcq_sheets") or []:
        if _nonempty_dataframe(m.get("df")):
            return False
    for om in sheets.get("optional_mcq_sheets") or []:
        if _nonempty_dataframe(om.get("df")):
            return False
    for ps in sheets.get("project_submission_sheets") or []:
        if _nonempty_dataframe(ps.get("df")):
            return False
    return True


def cohort2_action_center_optional_mcq_only_classification(classification):
    """Preview-time counterpart to cohort2_action_center_optional_mcq_only."""
    detected = (classification or {}).get("detected_by_module", {}) or {}
    if not detected.get("cohort2_optional_mcq"):
        return False
    for module in (
        "skilllab_submission",
        "codelab_submission",
        "lab_completion",
        "main_mcq",
        "optional_mcq",
        "project_submission",
    ):
        if detected.get(module):
            return False
    return True


def action_center_preview_has_recognized_imports(preview):
    """True if preview dict lists any importable subsheet besides the Skills profile sheet."""
    if not preview or not isinstance(preview, dict):
        return False
    if preview.get("submission_sheet_name"):
        return True
    if preview.get("codelab_sheet_name"):
        return True
    if preview.get("codelab_submission_sheets"):
        return True
    if preview.get("mcq_sheets"):
        return True
    c2 = preview.get("cohort2_optional_mcq_sheet") or {}
    if c2.get("sheet_name"):
        return True
    if preview.get("lab_completion_sheets"):
        return True
    if preview.get("main_mcq_sheets"):
        return True
    if preview.get("project_submission_sheets"):
        return True
    return False


def fill_action_center_secondary_preview(file_path, result, classification=None, cohort_id=None):
    """Populate submission / optional & main MCQ / lab / project / C2 optional keys on preview dict.

    Detection is fully delegated to classify_workbook_sheets() so a renumbered
    Action Center tab (e.g. "15.MCQ Track 1" instead of "16.MCQ Track 1") still
    maps to its module. The classification is reused if already computed.
    """
    if classification is None:
        classification = classify_workbook_sheets(file_path, cohort_id=cohort_id)
    detected = classification.get("detected_by_module", {}) or {}

    def _row_count(name):
        try:
            df = parse_excel_sheet(file_path, name)
            return len(df) if df is not None else 0
        except Exception:
            return 0

    # Skill Lab Submission
    try:
        for d in detected.get("skilllab_submission", []):
            name = d.get("sheet_name")
            if not name:
                continue
            result["submission_sheet_name"] = name
            result["submission_sheet_rows"] = _row_count(name)
            break
    except Exception:
        pass

    # Code Lab Submission (all matched tabs)
    result["codelab_submission_sheets"] = []
    try:
        for d in detected.get("codelab_submission", []):
            name = d.get("sheet_name")
            if not name:
                continue
            track = _codelab_sheet_track_number(name, d.get("track"))
            default_ps = _codelab_default_problem_statement(name, track)
            result["codelab_submission_sheets"].append({
                "track": track,
                "sheet_name": name,
                "rows": _row_count(name),
                "default_problem_statement": default_ps,
            })
        if result["codelab_submission_sheets"]:
            first = result["codelab_submission_sheets"][0]
            result["codelab_sheet_name"] = first["sheet_name"]
            result["codelab_sheet_rows"] = first["rows"]
    except Exception:
        pass

    # Optional MCQ sheets (Track 1–3)
    result["mcq_sheets"] = []
    try:
        om_rows = sorted(
            detected.get("optional_mcq", []),
            key=lambda r: (r.get("track") is None, r.get("track") or 0),
        )
        for d in om_rows:
            name = d.get("sheet_name")
            if not name:
                continue
            result["mcq_sheets"].append({
                "track": d.get("track"),
                "sheet_name": name,
                "rows": _row_count(name),
            })
    except Exception:
        pass

    # Cohort 2 Optional MCQ
    result["cohort2_optional_mcq_sheet"] = None
    try:
        c2_rows = detected.get("cohort2_optional_mcq", [])
        c2_name = c2_rows[0].get("sheet_name") if c2_rows else None
        if not c2_name and cohort_id is None:
            c2_name = find_cohort2_optional_mcq_sheet_name(classification.get("sheet_names") or [])
        if c2_name:
            result["cohort2_optional_mcq_sheet"] = {
                "sheet_name": c2_name,
                "rows": _row_count(c2_name),
            }
    except Exception:
        pass

    # Lab Completion sheets
    result["lab_completion_sheets"] = []
    try:
        for d in detected.get("lab_completion", []):
            name = d.get("sheet_name")
            if not name:
                continue
            result["lab_completion_sheets"].append({
                "lab": d.get("lab"),
                "track": d.get("track"),
                "sheet_name": name,
                "rows": _row_count(name),
            })
    except Exception:
        pass

    # Main MCQ sheets (Track 1–3) — prefix-agnostic
    result["main_mcq_sheets"] = []
    try:
        main_rows = sorted(
            detected.get("main_mcq", []),
            key=lambda r: (r.get("track") is None, r.get("track") or 0),
        )
        for d in main_rows:
            name = d.get("sheet_name")
            if not name:
                continue
            result["main_mcq_sheets"].append({
                "track": d.get("track"),
                "sheet_name": name,
                "rows": _row_count(name),
            })
    except Exception:
        pass

    # Project Submission Track 1–3
    result["project_submission_sheets"] = []
    try:
        ps_rows = sorted(
            detected.get("project_submission", []),
            key=lambda r: (r.get("track") is None, r.get("track") or 0),
        )
        for d in ps_rows:
            name = d.get("sheet_name")
            if not name:
                continue
            result["project_submission_sheets"].append({
                "track": d.get("track"),
                "sheet_name": name,
                "rows": _row_count(name),
            })
    except Exception:
        pass


def find_cohort2_optional_mcq_sheet_name(sheet_names):
    """
    Find the Cohort 2 Optional MCQ worksheet: first sheet whose name matches
    <number>.MCQ Optional (e.g. '14.MCQ Optional', '1. MCQ Optional'); letters case-insensitive.
    Returns the exact sheet name from the workbook, or None.
    """
    if not sheet_names:
        return None
    for raw in sheet_names:
        if raw is None:
            continue
        name = str(raw).replace("\u00a0", " ").strip()
        if _COHORT2_OPTIONAL_MCQ_SHEET_NAME_RE.fullmatch(name):
            return raw
    return None


def find_cohort2_optional_mcq_sheet_name_from_file(file_path):
    return find_cohort2_optional_mcq_sheet_name(get_sheet_names(file_path))


def find_sheet_by_substring(file_path, substring):
    """
    Find worksheet name that contains the given substring (case-insensitive).
    Returns the first matching sheet name, or None if none match.
    """
    try:
        if str(file_path).lower().endswith('.xlsx'):
            import openpyxl
            wb = openpyxl.load_workbook(file_path, read_only=True)
            for name in wb.sheetnames:
                if substring.lower() in name.lower():
                    wb.close()
                    return name
            wb.close()
        else:
            xl = pd.ExcelFile(file_path)
            for name in xl.sheet_names:
                if substring.lower() in name.lower():
                    return name
    except Exception:
        pass
    return None


def load_all_skillboost_sheets(file_path, cohort_id=None):
    """
    Open the workbook once and load all sheets needed for Action Center import.

    Sheet-to-module classification goes through classify_workbook_sheets() so the
    importer picks up renumbered Action Center tabs (e.g. "15.MCQ Track 1"
    instead of "16.MCQ Track 1") without code changes.

    Returns a dict: profile_sheet_name, profile_df, submission_sheet_name,
    submission_df, codelab_sheet_name, codelab_df,
    lab_completion_sheets    (list of {lab, track, sheet_name, df}),
    main_mcq_sheets          (list of {track, sheet_name, df}),
    optional_mcq_sheets      (list of {track, sheet_name, df}),
    project_submission_sheets(list of {track, sheet_name, df}),
    cohort2_optional_mcq_sheet (dict or None),
    plus the raw classification under 'classification' so callers can surface
    sheet_mapping / missing_critical_modules without re-walking the workbook.

    Missing sheets are omitted. Use this to avoid re-opening the same file many times.
    """
    result = {
        'profile_sheet_name': None,
        'profile_df': None,
        'submission_sheet_name': None,
        'submission_df': None,
        'codelab_sheet_name': None,
        'codelab_df': None,
        'codelab_submission_sheets': [],
        'lab_completion_sheets': [],
        'main_mcq_sheets': [],
        'optional_mcq_sheets': [],
        'cohort2_optional_mcq_sheet': None,
        'project_submission_sheets': [],
        'classification': None,
    }

    try:
        classification = classify_workbook_sheets(file_path, cohort_id=cohort_id)
        result['classification'] = classification
        detected = classification.get('detected_by_module', {}) or {}

        engine = 'openpyxl' if str(file_path).lower().endswith('.xlsx') else None
        xl = pd.ExcelFile(file_path, engine=engine)

        # Skill Lab profile (only the first matched tab is used as the profile)
        for det_row in detected.get('skillboost_profile', []):
            name = det_row.get('sheet_name')
            if not name:
                continue
            result['profile_sheet_name'] = name
            try:
                result['profile_df'] = pd.read_excel(xl, sheet_name=name)
            except Exception:
                result['profile_df'] = None
            break

        # Skill Lab Submission
        for det_row in detected.get('skilllab_submission', []):
            name = det_row.get('sheet_name')
            if not name:
                continue
            result['submission_sheet_name'] = name
            try:
                result['submission_df'] = pd.read_excel(xl, sheet_name=name)
            except Exception:
                result['submission_df'] = None
            break

        # Code Lab Submission (all matched tabs — Cohort 2: Student + Professional tracks)
        for det_row in detected.get('codelab_submission', []):
            name = det_row.get('sheet_name')
            if not name:
                continue
            track = _codelab_sheet_track_number(name, det_row.get('track'))
            default_ps = _codelab_default_problem_statement(name, track)
            try:
                df = pd.read_excel(xl, sheet_name=name)
            except Exception:
                df = None
            result['codelab_submission_sheets'].append({
                'track': track,
                'sheet_name': name,
                'df': df,
                'default_problem_statement': default_ps,
            })
        if result['codelab_submission_sheets']:
            first = result['codelab_submission_sheets'][0]
            result['codelab_sheet_name'] = first['sheet_name']
            result['codelab_df'] = first.get('df')

        # Lab Completion sheets (preserve registry order)
        for det_row in detected.get('lab_completion', []):
            name = det_row.get('sheet_name')
            if not name:
                continue
            try:
                df = pd.read_excel(xl, sheet_name=name)
            except Exception:
                df = None
            result['lab_completion_sheets'].append({
                'lab': det_row.get('lab'),
                'track': det_row.get('track'),
                'sheet_name': name,
                'df': df,
            })

        # Main MCQ sheets (sorted by track 1->3)
        main_rows = sorted(
            detected.get('main_mcq', []),
            key=lambda r: (r.get('track') is None, r.get('track') or 0),
        )
        for det_row in main_rows:
            name = det_row.get('sheet_name')
            if not name:
                continue
            try:
                df = pd.read_excel(xl, sheet_name=name)
            except Exception:
                df = None
            result['main_mcq_sheets'].append({
                'track': det_row.get('track'),
                'sheet_name': name,
                'df': df,
            })

        # Project Submission (sorted by track)
        ps_rows = sorted(
            detected.get('project_submission', []),
            key=lambda r: (r.get('track') is None, r.get('track') or 0),
        )
        for det_row in ps_rows:
            name = det_row.get('sheet_name')
            if not name:
                continue
            try:
                df = pd.read_excel(xl, sheet_name=name)
            except Exception:
                df = None
            result['project_submission_sheets'].append({
                'track': det_row.get('track'),
                'sheet_name': name,
                'df': df,
            })

        # Optional MCQ (sorted by track)
        om_rows = sorted(
            detected.get('optional_mcq', []),
            key=lambda r: (r.get('track') is None, r.get('track') or 0),
        )
        for det_row in om_rows:
            name = det_row.get('sheet_name')
            if not name:
                continue
            try:
                df = pd.read_excel(xl, sheet_name=name)
            except Exception:
                df = None
            result['optional_mcq_sheets'].append({
                'track': det_row.get('track'),
                'sheet_name': name,
                'df': df,
            })

        # Cohort 2 Optional MCQ (anchored "<digits>.MCQ Optional" pattern).
        # The detector is gated on cohort_id=2, but legacy callers may not pass
        # cohort_id; fall back to the historic name-only detection for Cohort 1
        # workflows so behaviour is unchanged when cohort_id is not provided.
        c2_rows = detected.get('cohort2_optional_mcq', [])
        c2_name = None
        if c2_rows:
            c2_name = c2_rows[0].get('sheet_name')
        elif cohort_id is None:
            c2_name = find_cohort2_optional_mcq_sheet_name(classification.get('sheet_names') or [])
        if c2_name:
            try:
                c2_df = pd.read_excel(xl, sheet_name=c2_name)
            except Exception:
                c2_df = None
            result['cohort2_optional_mcq_sheet'] = {
                'sheet_name': c2_name,
                'df': c2_df,
            }
    except Exception:
        pass
    return result


def load_cohort2_optional_mcq_sheet_only(file_path):
    """
    Load only the Cohort 2 Optional MCQ worksheet (e.g. "14.MCQ  Optional").
    Used when importing from Action Center without the Google Skills Boost profile sheet.
    """
    out = {'cohort2_optional_mcq_sheet': None}
    try:
        engine = 'openpyxl' if str(file_path).lower().endswith('.xlsx') else None
        xl = pd.ExcelFile(file_path, engine=engine)

        c2_name = find_cohort2_optional_mcq_sheet_name(xl.sheet_names)
        if c2_name:
            out['cohort2_optional_mcq_sheet'] = {
                'sheet_name': c2_name,
                'df': pd.read_excel(xl, sheet_name=c2_name),
            }
    except Exception:
        pass
    return out


def get_skillboost_preview(file_path, cohort_id=None):
    """
    Preview Skill Lab / Action Center XLSX: sheet list plus detected subsheets
    for import.

    Detection is fully driven by the SHEET_DETECTORS regex registry, so any tab
    whose name contains, e.g., "MCQ Track 1" is detected regardless of leading
    numbering ("16.MCQ Track 1", "15.MCQ Track 1", "MCQ Track 1", …).

    cohort_id=1 (or unset): Skills profile sheet is optional if the workbook
        contains other recognized subsheets (Main MCQ, Optional MCQ, …).
    cohort_id=2: without a Skills profile sheet, preview is Optional-MCQ–only if
        that tab exists; otherwise full subsheet mapping (split Action Center files).

    Response dict adds two top-level fields the UI uses to show a full mapping
    table and warn about missing critical modules:
        sheet_mapping             — list of {sheet_name, module, label, track,
                                    lab, status} (every tab in workbook order;
                                    status in detected/duplicate/unrecognised).
        missing_critical_modules  — list of {module, label, track, lab} that
                                    were expected for this cohort but did not
                                    match any tab.

    Existing per-module fields (main_mcq_sheets, mcq_sheets,
    cohort2_optional_mcq_sheet, lab_completion_sheets,
    project_submission_sheets, submission_sheet_name, submission_sheet_rows,
    detected_sheet_name, …) are derived from the same classification for
    backward compatibility.
    """
    sheet_names = get_sheet_names(file_path)
    sheet_count = len(sheet_names)
    classification = classify_workbook_sheets(file_path, cohort_id=cohort_id)
    detected = classification.get('detected_by_module', {}) or {}

    result = {
        'sheet_count': sheet_count,
        'sheet_names': sheet_names,
        'sheet_mapping': classification.get('sheet_mapping') or [],
        'missing_critical_modules': classification.get('missing_critical_modules') or [],
        'detected_sheet_name': None,
        'detected_sheet_rows': None,
        'columns': None,
        'error': None,
        'submission_sheet_name': None,
        'submission_sheet_rows': None,
        'cohort2_optional_import_only': False,
        'action_center_profile_sheet_missing': False,
    }
    if not sheet_names:
        result['error'] = 'Could not read any worksheets from the file.'
        return result

    profile_rows = detected.get('skillboost_profile', [])
    skill_sheet = profile_rows[0]['sheet_name'] if profile_rows else None

    # Cohort 2: Action Center export may omit the Google Skills sheet — Optional MCQ–only
    # preview when that is the sole recognised subsheet; otherwise full mapping (Code Lab,
    # Skill Lab submission, split exports, etc.).
    if cohort_id == 2 and not skill_sheet:
        if cohort2_action_center_optional_mcq_only_classification(classification):
            c2_rows = detected.get('cohort2_optional_mcq', [])
            c2_sheet = c2_rows[0]['sheet_name'] if c2_rows else None
            if c2_sheet:
                result['cohort2_optional_import_only'] = True
                try:
                    df = parse_excel_sheet(file_path, c2_sheet)
                    nrows = len(df) if df is not None else 0
                    result['detected_sheet_name'] = c2_sheet
                    result['detected_sheet_rows'] = nrows
                    result['columns'] = list(df.columns) if df is not None and len(df) > 0 else []
                    result['cohort2_optional_mcq_sheet'] = {'sheet_name': c2_sheet, 'rows': nrows}
                except Exception as e:
                    result['error'] = f'Error reading sheet "{c2_sheet}": {str(e)}'
                return result
        result['cohort2_optional_import_only'] = False
        result['action_center_profile_sheet_missing'] = True
        fill_action_center_secondary_preview(
            file_path, result, classification=classification, cohort_id=cohort_id,
        )
        if not action_center_preview_has_recognized_imports(result):
            result['error'] = ACTION_CENTER_NO_RECOGNIZED_SHEETS_MESSAGE
        return result

    result['cohort2_optional_import_only'] = False
    # Populate per-module preview keys from the same classification (no second pass).
    fill_action_center_secondary_preview(
        file_path, result, classification=classification, cohort_id=cohort_id,
    )

    if skill_sheet:
        result['detected_sheet_name'] = skill_sheet
        result['action_center_profile_sheet_missing'] = False
        try:
            df = parse_excel_sheet(file_path, skill_sheet)
            result['detected_sheet_rows'] = len(df) if df is not None else 0
            result['columns'] = list(df.columns) if df is not None and len(df) > 0 else []
        except Exception as e:
            result['error'] = f'Error reading profile sheet "{skill_sheet}": {str(e)}'
    else:
        result['action_center_profile_sheet_missing'] = True

    if cohort_id in (1, None) and not skill_sheet:
        if not action_center_preview_has_recognized_imports(result):
            result['error'] = ACTION_CENTER_NO_RECOGNIZED_SHEETS_MESSAGE

    return result


def parse_excel_sheet(file_path, sheet_name):
    """
    Parse a specific sheet from an Excel file into a DataFrame.
    """
    try:
        if str(file_path).lower().endswith('.xlsx'):
            df = pd.read_excel(file_path, engine='openpyxl', sheet_name=sheet_name)
        else:
            df = pd.read_excel(file_path, sheet_name=sheet_name)
        return df
    except Exception as e:
        raise Exception(f"Error parsing sheet '{sheet_name}': {str(e)}")


def _find_email_column(columns):
    """Return column name that best matches 'email' (case-insensitive, strip)."""
    for col in columns:
        if col is None:
            continue
        if str(col).strip().lower() == 'email':
            return col
    for col in columns:
        if col and 'email' in str(col).lower():
            return col
    return None


def _find_profile_link_column(columns):
    """
    Return column name that best matches profile link (contains 'profile', 'link', or 'skills').
    Prefer names like 'Share your Google Skills Boost public profile...' or 'Skillboost public view link'.
    """
    cols_with = []
    for col in columns:
        if col is None:
            continue
        s = str(col).lower()
        if 'profile' in s or 'link' in s or 'skills' in s:
            cols_with.append(col)
    if not cols_with:
        return None
    # Prefer one that has "profile" and ("link" or "skills")
    for c in cols_with:
        s = str(c).lower()
        if 'profile' in s and ('link' in s or 'skills' in s):
            return c
    return cols_with[0]


def parse_excel(file_path):
    """
    Parse Excel file and return DataFrame (all rows, no limit).
    
    Args:
        file_path: Path to Excel file
        
    Returns:
        pandas DataFrame
    """
    try:
        # Use openpyxl for .xlsx to read all rows reliably; no nrows limit
        if str(file_path).lower().endswith('.xlsx'):
            df = pd.read_excel(file_path, engine='openpyxl', sheet_name=0)
        else:
            df = pd.read_excel(file_path, sheet_name=0)
        return df
    except Exception as e:
        raise Exception(f"Error parsing Excel file: {str(e)}")


def get_db_fields():
    """Get list of database fields for UserPII table"""
    return [
        'registered_at',
        'organization_name',
        'class_stream',
        'domain',
        'designation',
        'name',
        'email',
        'mobile_number',
        'country',
        'state',
        'city',
        'date_of_birth',
        'gender',
        'occupation',
        'github_url',
        'linkedin_url',
        'utm_medium'
    ]


# Max string lengths for UserPII columns (match model) to avoid "data too long" errors
USERPII_STRING_MAX_LENGTHS = {
    'organization_name': 255,
    'class_stream': 255,
    'domain': 255,
    'designation': 255,
    'name': 255,
    'email': 255,
    'mobile_number': 50,
    'country': 100,
    'state': 100,
    'city': 100,
    'gender': 50,
    'occupation': 255,
    'github_url': 500,
    'linkedin_url': 500,
    'utm_medium': 255,
}


def truncate_record_strings(data):
    """Truncate string values in data to column max lengths. Modifies data in place."""
    for key, max_len in USERPII_STRING_MAX_LENGTHS.items():
        if key not in data or data[key] is None:
            continue
        val = data[key]
        if isinstance(val, str) and len(val) > max_len:
            data[key] = val[:max_len]


def apply_title_categories_to_record(data):
    """Populate sub_category and broad_category from designation (Cohort 2+)."""
    try:
        from flask import g
        if getattr(g, "cohort_id", None) not in (2, 3):
            return
    except RuntimeError:
        return
    designation = data.get("designation")
    if not designation:
        return
    from server.utils.title_map import get_title_categories

    sub, broad = get_title_categories(designation)
    data["sub_category"] = sub
    data["broad_category"] = broad


def normalize_field_name(name):
    """
    Normalize field name for matching
    - Convert to lowercase
    - Remove spaces, underscores, hyphens, slashes
    - Remove special characters
    """
    if not name:
        return ''
    normalized = str(name).lower()
    normalized = normalized.replace(' ', '')
    normalized = normalized.replace('_', '')
    normalized = normalized.replace('-', '')
    normalized = normalized.replace('.', '')
    normalized = normalized.replace('/', '')
    return normalized


def auto_map_fields(excel_columns):
    """
    Auto-map Excel columns to database fields
    
    Args:
        excel_columns: List of Excel column names
        
    Returns:
        Dictionary mapping Excel column names to DB field names
    """
    db_fields = get_db_fields()
    mapping = {}

    # Columns to skip (do not map to any DB field). Normalized names.
    skip_columns_normalized = {
        'collegeschoolstate',   # College/School State
        'collegeschoolcity',    # College/School city
        'profilename',          # Profile Name
        'innovator',
        'utmsource',
        'utmcampaign',
        'utmterm',
        'utmcontent',
        'portfolio',
        'foundedinstartupsize',
        'degreepassoutyear',
        'whatsapp',
        'role',
    }

    # Explicit Excel header -> DB field (normalized header -> db field)
    explicit_column_map = {
        'timestamp': 'registered_at',
        'collegeschoolcompanystartupname': 'organization_name',  # College/School/Company/Startup Name
        'fullname': 'name',
        'name': 'name',
        'designationyearofexp': 'designation',
        'classstream': 'class_stream',
    }

    # Create normalized versions for matching
    normalized_db_fields = {normalize_field_name(field): field for field in db_fields}

    for excel_col in excel_columns:
        normalized_excel = normalize_field_name(excel_col)

        # Skip columns (do not map)
        if normalized_excel in skip_columns_normalized:
            mapping[excel_col] = None
            continue

        # Explicit mappings (Timestamp -> registered_at, College/School/Company/Startup Name -> organization_name)
        if normalized_excel in explicit_column_map:
            mapping[excel_col] = explicit_column_map[normalized_excel]
            continue

        # Try exact match first
        if normalized_excel in normalized_db_fields:
            mapping[excel_col] = normalized_db_fields[normalized_excel]
        else:
            # Try partial matches
            matched = None
            for norm_db, db_field in normalized_db_fields.items():
                if normalized_excel in norm_db or norm_db in normalized_excel:
                    matched = db_field
                    break

            # Common aliases (do not override explicit/skip)
            if not matched:
                alias_map = {
                    'org': 'organization_name',
                    'orgname': 'organization_name',
                    'company': 'organization_name',
                    'stream': 'class_stream',
                    'class': 'class_stream',
                    'dob': 'date_of_birth',
                    'birthdate': 'date_of_birth',
                    'phone': 'mobile_number',
                    'mobile': 'mobile_number',
                    'github': 'github_url',
                    'linkedin': 'linkedin_url',
                    'utmmedium': 'utm_medium',
                    'utm_medium': 'utm_medium',
                    'title': 'designation',
                    'jobtitle': 'designation',
                }

                for alias, db_field in alias_map.items():
                    if alias in normalized_excel:
                        matched = db_field
                        break

            mapping[excel_col] = matched if matched else None

    return mapping


def parse_date(date_value):
    """Parse date value from various formats"""
    if pd.isna(date_value):
        return None
    
    if isinstance(date_value, datetime):
        return date_value.date()
    
    if isinstance(date_value, str):
        try:
            return pd.to_datetime(date_value).date()
        except:
            return None
    
    return None


def import_data(df, mappings, mode='create', progress_callback=None):
    """
    Import data from DataFrame to database with batch processing for performance
    
    Args:
        df: pandas DataFrame
        mappings: Dictionary mapping Excel columns to DB fields
        mode: 'create', 'create_update', or 'update_only'
        progress_callback: Optional callable(created, updated, skipped) called periodically during import
        
    Returns:
        Dictionary with import summary
    """
    from server.models import db

    IM = _import_models()
    UserPII = IM.UserPII

    total_rows = len(df)
    created = 0
    updated = 0
    skipped = 0
    errors = []
    
    def report_progress():
        if progress_callback:
            try:
                progress_callback(created, updated, skipped)
            except Exception:
                pass
    
    # Filter out None mappings (unmapped columns)
    active_mappings = {k: v for k, v in mappings.items() if v is not None}
    
    # Batch size for database operations (commit every N records)
    BATCH_SIZE = 1000
    PROGRESS_INTERVAL = 100  # Report progress every N rows
    
    # Fetch existing emails in chunks (all modes) to avoid slow single query and OOM on large DBs
    existing_emails_set = set()
    CHUNK_SIZE = 20000
    try:
        offset = 0
        while True:
            chunk = db.session.query(UserPII.email).limit(CHUNK_SIZE).offset(offset).all()
            if not chunk:
                break
            existing_emails_set.update(e[0] for e in chunk if e[0])
            if len(chunk) < CHUNK_SIZE:
                break
            offset += CHUNK_SIZE
    except Exception:
        existing_emails_set = set()
    
    # Prepare data for batch insert
    records_to_insert = []
    records_to_update = []
    
    # Track emails we've already processed in this batch to avoid duplicates within the same file
    processed_emails_in_batch = set()
    last_reported = 0
    
    for index, row in df.iterrows():
        try:
            # Build data dictionary from mappings
            data = {}
            for excel_col, db_field in active_mappings.items():
                if excel_col in row:
                    value = row[excel_col]
                    
                    # Handle NaN values
                    if pd.isna(value):
                        value = None
                    # Handle date fields
                    elif db_field == 'date_of_birth':
                        value = parse_date(value)
                    elif db_field == 'registered_at':
                        value = parse_date(value)
                        if value:
                            # Convert date to datetime
                            from datetime import datetime
                            value = datetime.combine(value, datetime.min.time())
                    else:
                        # Convert to string for other fields
                        value = str(value).strip() if value else None
                    
                    data[db_field] = value
            
            truncate_record_strings(data)
            apply_title_categories_to_record(data)
            
            # Email is required
            if not data.get('email'):
                skipped += 1
                if len(errors) < 100:  # Limit error messages
                    errors.append(f"Row {index + 2}: Missing email")
                continue
            
            email = data['email']
            
            # Check for duplicate emails within the same file
            if email in processed_emails_in_batch:
                skipped += 1
                if len(errors) < 100:
                    errors.append(f"Row {index + 2}: Duplicate email in file")
                continue
            
            processed_emails_in_batch.add(email)
            
            if mode == 'create':
                if email in existing_emails_set:
                    skipped += 1
                    if len(errors) < 100:
                        errors.append(f"Row {index + 2}: Email already exists")
                    continue
                
                # Add to batch insert
                records_to_insert.append(data)
                created += 1
                
            elif mode == 'create_update':
                if email in existing_emails_set:
                    # Add to batch update
                    records_to_update.append((email, data))
                    updated += 1
                else:
                    # Add to batch insert (email doesn't exist in DB)
                    records_to_insert.append(data)
                    created += 1
                    # Also add to processed set to track what we're inserting
                    existing_emails_set.add(email)  # Track newly inserted emails
                    
            elif mode == 'update_only':
                if email in existing_emails_set:
                    # Add to batch update
                    records_to_update.append((email, data))
                    updated += 1
                else:
                    skipped += 1
                    if len(errors) < 100:
                        errors.append(f"Row {index + 2}: Email not found")
                    continue
            
            # Batch commit for inserts (create_update: on conflict process one-by-one so batch isn't lost)
            if len(records_to_insert) >= BATCH_SIZE:
                try:
                    if mode == 'create_update':
                        # Try batch insert first; on unique constraint rollback and process one-by-one
                        batch_ok = True
                        first_conflict_email = None
                        for record in records_to_insert:
                            try:
                                user = UserPII(**record)
                                db.session.add(user)
                                db.session.flush()
                            except Exception as insert_error:
                                db.session.rollback()
                                batch_ok = False
                                first_conflict_email = record.get('email')
                                if 'unique' in str(insert_error).lower() or 'duplicate' in str(insert_error).lower():
                                    existing_user = UserPII.query.filter_by(email=first_conflict_email).first()
                                    if existing_user:
                                        for key, value in record.items():
                                            setattr(existing_user, key, value)
                                        db.session.commit()
                                        updated += 1
                                        created -= 1
                                    else:
                                        if len(errors) < 100:
                                            errors.append(f"Row: {first_conflict_email} - {str(insert_error)}")
                                else:
                                    if len(errors) < 100:
                                        errors.append(f"Insert error {first_conflict_email}: {str(insert_error)}")
                                break
                        if batch_ok:
                            db.session.commit()
                            records_to_insert = []
                        else:
                            # Process remaining records one-by-one (skip already-handled conflict)
                            for record in records_to_insert:
                                if record.get('email') == first_conflict_email:
                                    continue
                                try:
                                    user = UserPII(**record)
                                    db.session.add(user)
                                    db.session.flush()
                                    db.session.commit()
                                    created += 1
                                except Exception as insert_error:
                                    db.session.rollback()
                                    if 'unique' in str(insert_error).lower() or 'duplicate' in str(insert_error).lower():
                                        existing_user = UserPII.query.filter_by(email=record.get('email')).first()
                                        if existing_user:
                                            for key, value in record.items():
                                                setattr(existing_user, key, value)
                                            db.session.commit()
                                            updated += 1
                                            created -= 1
                                        else:
                                            if len(errors) < 100:
                                                errors.append(f"Row: {record.get('email')} - {str(insert_error)}")
                                    else:
                                        if len(errors) < 100:
                                            errors.append(f"Insert error {record.get('email')}: {str(insert_error)}")
                            records_to_insert = []
                    else:
                        for record in records_to_insert:
                            user = UserPII(**record)
                            db.session.add(user)
                        db.session.commit()
                        records_to_insert = []
                except Exception as e:
                    db.session.rollback()
                    if len(errors) < 100:
                        errors.append(f"Batch insert error: {str(e)}")
            
            # Batch commit for updates: one query to fetch all users in batch, then update in memory
            if len(records_to_update) >= BATCH_SIZE:
                try:
                    emails_batch = [e for e, _ in records_to_update]
                    users_map = {u.email: u for u in UserPII.query.filter(UserPII.email.in_(emails_batch)).all()}
                    for update_email, update_data in records_to_update:
                        existing_user = users_map.get(update_email)
                        if existing_user:
                            for key, value in update_data.items():
                                setattr(existing_user, key, value)
                    db.session.commit()
                    records_to_update = []
                except Exception as e:
                    db.session.rollback()
                    if len(errors) < 100:
                        errors.append(f"Batch update error: {str(e)}")
            
            # Report progress periodically
            processed = created + updated + skipped
            if processed - last_reported >= PROGRESS_INTERVAL:
                last_reported = processed
                report_progress()
            
        except Exception as e:
            skipped += 1
            if len(errors) < 100:
                errors.append(f"Row {index + 2}: {str(e)}")
    
    # Commit remaining records (create_update: one-by-one so no batch lost on unique constraint)
    try:
        if records_to_insert:
            if mode == 'create_update':
                for record in records_to_insert:
                    try:
                        user = UserPII(**record)
                        db.session.add(user)
                        db.session.flush()
                        db.session.commit()
                    except Exception as insert_error:
                        db.session.rollback()
                        if 'unique' in str(insert_error).lower() or 'duplicate' in str(insert_error).lower():
                            existing_user = UserPII.query.filter_by(email=record.get('email')).first()
                            if existing_user:
                                for key, value in record.items():
                                    setattr(existing_user, key, value)
                                db.session.commit()
                                updated += 1
                                created -= 1
                            else:
                                if len(errors) < 100:
                                    errors.append(f"Row: {record.get('email')} - {str(insert_error)}")
                        else:
                            if len(errors) < 100:
                                errors.append(f"Insert error {record.get('email')}: {str(insert_error)}")
            else:
                try:
                    for record in records_to_insert:
                        user = UserPII(**record)
                        db.session.add(user)
                    db.session.commit()
                except (DataError, IntegrityError):
                    db.session.rollback()
                    for record in records_to_insert:
                        try:
                            user = UserPII(**record)
                            db.session.add(user)
                            db.session.commit()
                        except Exception as insert_error:
                            db.session.rollback()
                            if len(errors) < 100:
                                errors.append(f"Row: {record.get('email', '')} - {str(insert_error)[:150]}")
        
        if records_to_update:
            emails_batch = [e for e, _ in records_to_update]
            try:
                users_map = {u.email: u for u in UserPII.query.filter(UserPII.email.in_(emails_batch)).all()}
                for update_email, update_data in records_to_update:
                    existing_user = users_map.get(update_email)
                    if existing_user:
                        for key, value in update_data.items():
                            setattr(existing_user, key, value)
                db.session.commit()
            except (DataError, IntegrityError) as e:
                db.session.rollback()
                if len(errors) < 100:
                    errors.append(f"Batch update failed: {str(e)[:150]}")
    except (DataError, IntegrityError) as e:
        db.session.rollback()
        err_msg = str(e).split('\n')[0][:200] if str(e) else 'Database constraint or data length error'
        if 'truncat' in str(e).lower() or 'too long' in str(e).lower() or 'overflow' in str(e).lower():
            err_msg = 'One or more values exceed column max length (e.g. organization_name 255 chars). Data is truncated automatically; if this persists, check the row data.'
        raise Exception(f"Database error: {err_msg}")
    except Exception as e:
        db.session.rollback()
        raise Exception(f"Database error: {str(e)}")
    
    report_progress()  # Final progress
    return {
        'total_rows': total_rows,
        'created': created,
        'updated': updated,
        'skipped': skipped,
        'errors': errors[:100]  # Limit errors to first 100
    }


def import_data_injected(df, mappings, mode='create', progress_callback=None):
    """
    Import data into user_pii_injected table. Same logic as import_data but targets UserPIIInjected.
    Does not run recalculate_bob_match.
    """
    from server.models import db

    IM = _import_models()
    UserPIIInjected = IM.UserPIIInjected

    total_rows = len(df)
    created = 0
    updated = 0
    skipped = 0
    errors = []

    def report_progress():
        if progress_callback:
            try:
                progress_callback(created, updated, skipped)
            except Exception:
                pass

    active_mappings = {k: v for k, v in mappings.items() if v is not None}
    BATCH_SIZE = 1000
    PROGRESS_INTERVAL = 100

    # Only default registered_at to now() when no column is mapped to it
    has_registered_at_mapping = 'registered_at' in active_mappings.values()

    existing_emails_set = set()
    CHUNK_SIZE = 20000
    try:
        offset = 0
        while True:
            chunk = db.session.query(UserPIIInjected.email).limit(CHUNK_SIZE).offset(offset).all()
            if not chunk:
                break
            existing_emails_set.update(e[0] for e in chunk if e[0])
            if len(chunk) < CHUNK_SIZE:
                break
            offset += CHUNK_SIZE
    except Exception:
        existing_emails_set = set()

    records_to_insert = []
    records_to_update = []
    processed_emails_in_batch = set()
    last_reported = 0

    for index, row in df.iterrows():
        try:
            data = {}
            for excel_col, db_field in active_mappings.items():
                if excel_col in row:
                    value = row[excel_col]
                    if pd.isna(value):
                        value = None
                    elif db_field == 'date_of_birth':
                        value = parse_date(value)
                    elif db_field == 'registered_at':
                        value = parse_date(value)
                        if value:
                            from datetime import datetime
                            value = datetime.combine(value, datetime.min.time())
                    else:
                        value = str(value).strip() if value else None
                    data[db_field] = value

            truncate_record_strings(data)
            apply_title_categories_to_record(data)
            if not data.get('email'):
                skipped += 1
                if len(errors) < 100:
                    errors.append(f"Row {index + 2}: Missing email")
                continue
            if not (data.get('name') and str(data.get('name')).strip()):
                skipped += 1
                if len(errors) < 100:
                    errors.append(f"Row {index + 2}: Missing name")
                continue
            # Only default to today when no column was mapped to registered_at
            if not data.get('registered_at') and not has_registered_at_mapping:
                from datetime import datetime as _dt
                data['registered_at'] = _dt.now()

            email = data['email']
            if email in processed_emails_in_batch:
                skipped += 1
                if len(errors) < 100:
                    errors.append(f"Row {index + 2}: Duplicate email in file")
                continue
            processed_emails_in_batch.add(email)

            if mode == 'create':
                if email in existing_emails_set:
                    skipped += 1
                    if len(errors) < 100:
                        errors.append(f"Row {index + 2}: Email already exists")
                    continue
                records_to_insert.append(data)
                created += 1
            elif mode == 'create_update':
                if email in existing_emails_set:
                    records_to_update.append((email, data))
                    updated += 1
                else:
                    records_to_insert.append(data)
                    created += 1
                    existing_emails_set.add(email)
            elif mode == 'update_only':
                if email in existing_emails_set:
                    records_to_update.append((email, data))
                    updated += 1
                else:
                    skipped += 1
                    if len(errors) < 100:
                        errors.append(f"Row {index + 2}: Email not found")
                    continue

            if len(records_to_insert) >= BATCH_SIZE:
                try:
                    if mode == 'create_update':
                        batch_ok = True
                        first_conflict_email = None
                        for record in records_to_insert:
                            try:
                                user = UserPIIInjected(**record)
                                db.session.add(user)
                                db.session.flush()
                            except Exception as insert_error:
                                db.session.rollback()
                                batch_ok = False
                                first_conflict_email = record.get('email')
                                if 'unique' in str(insert_error).lower() or 'duplicate' in str(insert_error).lower():
                                    existing_user = UserPIIInjected.query.filter_by(email=first_conflict_email).first()
                                    if existing_user:
                                        for key, value in record.items():
                                            setattr(existing_user, key, value)
                                        db.session.commit()
                                        updated += 1
                                        created -= 1
                                    else:
                                        if len(errors) < 100:
                                            errors.append(f"Row: {first_conflict_email} - {str(insert_error)}")
                                else:
                                    if len(errors) < 100:
                                        errors.append(f"Insert error {first_conflict_email}: {str(insert_error)}")
                                break
                        if batch_ok:
                            db.session.commit()
                            records_to_insert = []
                        else:
                            for record in records_to_insert:
                                if record.get('email') == first_conflict_email:
                                    continue
                                try:
                                    user = UserPIIInjected(**record)
                                    db.session.add(user)
                                    db.session.flush()
                                    db.session.commit()
                                    created += 1
                                except Exception as insert_error:
                                    db.session.rollback()
                                    if 'unique' in str(insert_error).lower() or 'duplicate' in str(insert_error).lower():
                                        existing_user = UserPIIInjected.query.filter_by(email=record.get('email')).first()
                                        if existing_user:
                                            for key, value in record.items():
                                                setattr(existing_user, key, value)
                                            db.session.commit()
                                            updated += 1
                                            created -= 1
                                        else:
                                            if len(errors) < 100:
                                                errors.append(f"Row: {record.get('email')} - {str(insert_error)}")
                                    else:
                                        if len(errors) < 100:
                                            errors.append(f"Insert error {record.get('email')}: {str(insert_error)}")
                            records_to_insert = []
                    else:
                        for record in records_to_insert:
                            user = UserPIIInjected(**record)
                            db.session.add(user)
                        db.session.commit()
                        records_to_insert = []
                except Exception as e:
                    db.session.rollback()
                    if len(errors) < 100:
                        errors.append(f"Batch insert error: {str(e)}")

            if len(records_to_update) >= BATCH_SIZE:
                try:
                    emails_batch = [e for e, _ in records_to_update]
                    users_map = {u.email: u for u in UserPIIInjected.query.filter(UserPIIInjected.email.in_(emails_batch)).all()}
                    for update_email, update_data in records_to_update:
                        existing_user = users_map.get(update_email)
                        if existing_user:
                            for key, value in update_data.items():
                                setattr(existing_user, key, value)
                    db.session.commit()
                    records_to_update = []
                except Exception as e:
                    db.session.rollback()
                    if len(errors) < 100:
                        errors.append(f"Batch update error: {str(e)}")

            processed = created + updated + skipped
            if processed - last_reported >= PROGRESS_INTERVAL:
                last_reported = processed
                report_progress()
        except Exception as e:
            skipped += 1
            if len(errors) < 100:
                errors.append(f"Row {index + 2}: {str(e)}")

    try:
        if records_to_insert:
            if mode == 'create_update':
                for record in records_to_insert:
                    try:
                        user = UserPIIInjected(**record)
                        db.session.add(user)
                        db.session.flush()
                        db.session.commit()
                    except Exception as insert_error:
                        db.session.rollback()
                        if 'unique' in str(insert_error).lower() or 'duplicate' in str(insert_error).lower():
                            existing_user = UserPIIInjected.query.filter_by(email=record.get('email')).first()
                            if existing_user:
                                for key, value in record.items():
                                    setattr(existing_user, key, value)
                                db.session.commit()
                                updated += 1
                                created -= 1
                            else:
                                if len(errors) < 100:
                                    errors.append(f"Row: {record.get('email')} - {str(insert_error)}")
                        else:
                            if len(errors) < 100:
                                errors.append(f"Insert error {record.get('email')}: {str(insert_error)}")
            else:
                try:
                    for record in records_to_insert:
                        user = UserPIIInjected(**record)
                        db.session.add(user)
                    db.session.commit()
                except (DataError, IntegrityError):
                    db.session.rollback()
                    for record in records_to_insert:
                        try:
                            user = UserPIIInjected(**record)
                            db.session.add(user)
                            db.session.commit()
                        except Exception as insert_error:
                            db.session.rollback()
                            if len(errors) < 100:
                                errors.append(f"Row: {record.get('email', '')} - {str(insert_error)[:150]}")

        if records_to_update:
            emails_batch = [e for e, _ in records_to_update]
            try:
                users_map = {u.email: u for u in UserPIIInjected.query.filter(UserPIIInjected.email.in_(emails_batch)).all()}
                for update_email, update_data in records_to_update:
                    existing_user = users_map.get(update_email)
                    if existing_user:
                        for key, value in update_data.items():
                            setattr(existing_user, key, value)
                db.session.commit()
            except (DataError, IntegrityError) as e:
                db.session.rollback()
                if len(errors) < 100:
                    errors.append(f"Batch update failed: {str(e)[:150]}")
    except (DataError, IntegrityError) as e:
        db.session.rollback()
        err_msg = str(e).split('\n')[0][:200] if str(e) else 'Database constraint or data length error'
        if 'truncat' in str(e).lower() or 'too long' in str(e).lower() or 'overflow' in str(e).lower():
            err_msg = 'One or more values exceed column max length.'
        raise Exception(err_msg)
    except Exception as e:
        db.session.rollback()
        raise Exception(f"Database error: {str(e)}")

    report_progress()
    return {
        'total_rows': total_rows,
        'created': created,
        'updated': updated,
        'skipped': skipped,
        'errors': errors[:100]
    }


# ─────────────────────────────────────────────────────────────────────────────
# Per-row error classification + structured error reporting
#
# Every importer wraps each row in db.session.begin_nested() (a SAVEPOINT).
# When a row raises, _classify_row_error converts the exception into a stable
# reason_code so the result UI can group similar issues. The structured row
# dicts also ride into the import response under per_sheet_errors so users
# get a concrete view of what failed and can download a CSV.
# ─────────────────────────────────────────────────────────────────────────────

ROW_ERROR_REASONS = (
    'missing_email',
    'invalid_email',
    'fk_user_pii_missing',
    'unique_constraint_violation',
    'value_too_long',
    'data_type_error',
    'integrity_error',
    'other',
)


def _classify_row_error(index, exc, raw_email=None, *, reason_code=None,
                        reason_message=None):
    """Return a structured error dict for one failed row.

    index   : pandas row index (0-based); converted to a 1-based sheet row
              (header is row 1, first data row is row 2).
    exc     : the exception raised, OR None when the caller already knows
              the reason (e.g. missing_email pre-check).
    """
    row_no = (int(index) + 2) if index is not None else None

    if reason_code is None:
        reason_code = 'other'
        try:
            from sqlalchemy.exc import IntegrityError, DataError
        except Exception:
            IntegrityError = DataError = None  # type: ignore

        msg = (str(exc) if exc is not None else '').lower()

        if IntegrityError is not None and isinstance(exc, IntegrityError):
            if 'foreign key' in msg or 'foreignkeyviolation' in msg:
                reason_code = 'fk_user_pii_missing'
            elif 'unique' in msg or 'duplicate' in msg:
                reason_code = 'unique_constraint_violation'
            else:
                reason_code = 'integrity_error'
        elif DataError is not None and isinstance(exc, DataError):
            if 'too long' in msg or 'value too long' in msg or 'truncat' in msg:
                reason_code = 'value_too_long'
            else:
                reason_code = 'data_type_error'

    if reason_message is None:
        reason_message = (str(exc) if exc is not None else '')[:300]

    return {
        'row': row_no,
        'reason_code': reason_code,
        'reason_message': reason_message,
        'raw_email': (raw_email or '')[:255],
    }


# Legacy `errors` strings power the import UI warning list; omit expected outcomes.
_INFORMATIONAL_IMPORT_REASON_CODES = frozenset({'profile_verified_no_update'})


def _record_row_error(errors, rows_errors, error_dict, max_errors=100,
                      max_rows_errors=2000):
    """Append a structured row-error to `rows_errors` and optionally the legacy
    `errors` text list (capped). Informational reason codes (e.g. verified profile
    left unchanged) are kept in `rows_errors` / CSV / per-sheet breakdown only,
    not in `errors`, so the UI does not treat them as failures.
    """
    if rows_errors is not None and len(rows_errors) < max_rows_errors:
        rows_errors.append(error_dict)
    code = error_dict.get('reason_code')
    if code in _INFORMATIONAL_IMPORT_REASON_CODES:
        return
    if errors is not None and len(errors) < max_errors:
        row = error_dict.get('row')
        msg = error_dict.get('reason_message') or error_dict.get('reason_code') or ''
        errors.append(f"Row {row}: [{code}] {msg}")


def _ensure_user_pii_for_leaders(df, email_col, *, name_col=None,
                                 phone_col=None, organization_col=None):
    """For every distinct leader_email in `df` that's not already in user_pii,
    insert a minimal UserPII row (name/mobile_number filled when available)
    inside its own SAVEPOINT so a single bad row never aborts the rest.

    Returns dict {pii_auto_created, pii_auto_skipped, by_email_lower}, where
    by_email_lower maps lower(email) -> UserPII (every row that now exists).
    """
    from sqlalchemy import func
    from server.models import db

    IM = _import_models()
    UserPII = IM.UserPII

    out = {
        'pii_auto_created': 0,
        'pii_auto_skipped': 0,
        'by_email_lower': {},
    }
    if df is None or len(df) == 0 or not email_col or email_col not in df.columns:
        return out

    distinct_lowers = []
    seen = set()
    for v in df[email_col].dropna():
        e = str(v).strip().lower()
        if not e or '@' not in e or e in seen:
            continue
        seen.add(e)
        distinct_lowers.append(e)
    if not distinct_lowers:
        return out

    existing_by_lower = {}
    try:
        for chunk in _chunk_list(distinct_lowers, IMPORT_EMAIL_IN_CHUNK):
            if not chunk:
                continue
            for p in UserPII.query.filter(func.lower(UserPII.email).in_(chunk)).all():
                el = (p.email or '').strip().lower()
                if el:
                    existing_by_lower[el] = p
    except Exception:
        pass
    out['by_email_lower'] = dict(existing_by_lower)

    missing_lowers = [e for e in distinct_lowers if e not in existing_by_lower]
    if not missing_lowers:
        return out

    name_lookup = {}
    phone_lookup = {}
    org_lookup = {}
    try:
        for _, row in df.iterrows():
            try:
                ev = row.get(email_col)
                if ev is None or pd.isna(ev):
                    continue
                el = str(ev).strip().lower()
                if not el or el in existing_by_lower:
                    continue
                if name_col and name_col in df.columns and el not in name_lookup:
                    nv = row.get(name_col)
                    if nv is not None and not pd.isna(nv):
                        name_lookup[el] = str(nv).strip()[:255]
                if phone_col and phone_col in df.columns and el not in phone_lookup:
                    pv = row.get(phone_col)
                    if pv is not None and not pd.isna(pv):
                        phone_lookup[el] = str(pv).strip()[:50]
                if organization_col and organization_col in df.columns and el not in org_lookup:
                    ov = row.get(organization_col)
                    if ov is not None and not pd.isna(ov):
                        org_lookup[el] = str(ov).strip()[:255]
            except Exception:
                continue
    except Exception:
        pass

    for el in missing_lowers:
        try:
            with db.session.begin_nested():
                u = UserPII(
                    email=el,
                    name=name_lookup.get(el),
                    mobile_number=phone_lookup.get(el),
                    organization_name=org_lookup.get(el),
                )
                db.session.add(u)
            existing_by_lower[el] = u
            out['pii_auto_created'] += 1
        except Exception:
            out['pii_auto_skipped'] += 1

    out['by_email_lower'] = existing_by_lower
    return out


def import_skillboost_profile(df, email_col, profile_link_col, progress_callback=None):
    """
    Import Skill Lab / Google Skills Boost profiles from a DataFrame into skillboost_profile.
    - Create new rows for (email, link) that don't exist.
    - Update only when existing row has valid = FALSE; never overwrite valid = TRUE.
    - Email: required; strip, lowercase. Link: optional; if missing/empty store empty string (import row).
    - Skips rows with missing email, duplicate (email, link) later in the file,
      or an existing row that is already verified (valid = TRUE; never overwritten).
      Those verified unchanged rows are informational only (rows_errors / counts), not legacy ``errors``.
    - All emails are imported (no skip for email not in user_pii).

    Large files: preloads existing rows only for emails present in the sheet; batched commits;
    one savepoint per row so a single bad row does not lose the whole batch.
    """
    from server.models import db

    IM = _import_models()
    SkillboostProfile = IM.SkillboostProfile

    total_rows = len(df)
    created = 0
    updated = 0
    skipped = 0
    errors = []
    rows_errors = []
    skip_reason_counts = Counter()

    if not email_col or email_col not in df.columns:
        raise Exception("Email column not found in sheet")
    if not profile_link_col or profile_link_col not in df.columns:
        raise Exception("Profile link column not found in sheet")

    emails_for_lookup = set()
    try:
        ser = df[email_col].dropna().astype(str).str.strip().str.lower()
        emails_for_lookup = set(ser[ser.str.len() > 0])
    except Exception:
        pass

    existing = {}
    try:
        for chunk in _chunk_list(emails_for_lookup, IMPORT_EMAIL_IN_CHUNK):
            if not chunk:
                continue
            for sp_row in SkillboostProfile.query.filter(SkillboostProfile.email.in_(chunk)).all():
                key = (sp_row.email, sp_row.google_cloud_skills_boost_profile_link or '')
                existing[key] = sp_row
    except Exception:
        pass

    inserted_this_run = set()

    email_idx = df.columns.get_loc(email_col)
    link_idx = df.columns.get_loc(profile_link_col)
    email_series = df.iloc[:, email_idx]
    link_series = df.iloc[:, link_idx]

    for row_pos in range(total_rows):
        try:
            email_val = email_series.iat[row_pos]
            link_val = link_series.iat[row_pos]
            if pd.isna(email_val):
                skipped += 1
                skip_reason_counts['missing_email'] += 1
                _record_row_error(errors, rows_errors, _classify_row_error(
                    row_pos, None, raw_email='', reason_code='missing_email',
                    reason_message='Missing email',
                ))
                continue
            email = str(email_val).strip().lower()
            if not email:
                skipped += 1
                skip_reason_counts['missing_email'] += 1
                _record_row_error(errors, rows_errors, _classify_row_error(
                    row_pos, None, raw_email='', reason_code='missing_email',
                    reason_message='Missing email',
                ))
                continue
            if pd.isna(link_val):
                link = ''
            else:
                link = str(link_val).strip()
            if len(link) > 1024:
                link = link[:1024]
            key = (email, link)
            if key in existing:
                rec = existing[key]
                if rec.valid:
                    skipped += 1
                    skip_reason_counts['profile_verified_no_update'] += 1
                    continue
                if key in inserted_this_run:
                    skipped += 1
                    skip_reason_counts['duplicate_row_in_workbook'] += 1
                    _record_row_error(errors, rows_errors, _classify_row_error(
                        row_pos, None, raw_email=email,
                        reason_code='duplicate_row_in_workbook',
                        reason_message=(
                            'Duplicate email + profile link after an earlier row in '
                            'this file; only the first occurrence is kept.'
                        ),
                    ))
                    continue
                with db.session.begin_nested():
                    rec.updated_at = datetime.utcnow()
                updated += 1
            else:
                with db.session.begin_nested():
                    rec = SkillboostProfile(
                        email=email,
                        google_cloud_skills_boost_profile_link=link,
                        valid=False,
                        remarks=None,
                    )
                    db.session.add(rec)
                existing[key] = rec
                inserted_this_run.add(key)
                created += 1

            if progress_callback and (row_pos % 250 == 0 or row_pos == total_rows - 1):
                try:
                    progress_callback(created, updated, skipped)
                except Exception:
                    pass
        except Exception as e:
            skipped += 1
            err_d = _classify_row_error(
                row_pos, e, raw_email=str(email_val or '')[:255],
            )
            skip_reason_counts[err_d.get('reason_code') or 'other'] += 1
            _record_row_error(errors, rows_errors, err_d)

    try:
        db.session.commit()
    except Exception:
        db.session.rollback()
        return {
            'total_rows': total_rows,
            'created': created,
            'updated': updated,
            'skipped': skipped,
            'errors': errors[:100],
            'rows_errors': rows_errors,
            'skip_reason_counts': dict(skip_reason_counts),
        }

    try:
        from server.utils.skillboost_profile_reconcile import reconcile_skillboost_for_emails

        reconcile_skillboost_for_emails(emails_for_lookup or [])
    except Exception:
        try:
            from flask import current_app

            current_app.logger.warning(
                'skillboost_profile reconcile after import failed',
                exc_info=True,
            )
        except Exception:
            pass

    return {
        'total_rows': total_rows,
        'created': created,
        'updated': updated,
        'skipped': skipped,
        'errors': errors[:100],
        'rows_errors': rows_errors,
        'skip_reason_counts': dict(skip_reason_counts),
    }


# ─────────────────────────────────────────────────────────────────────────────
# Skill Lab Submission import (one DB row per sheet row; dedupe on re-import)
# ─────────────────────────────────────────────────────────────────────────────

# Column name mapping: normalised XLSX header -> model attribute
_SUBMISSION_COL_MAP = {
    'team name': 'team_name',
    'team_name': 'team_name',
    'leader name': 'leader_name',
    'leader_name': 'leader_name',
    'leader email': 'leader_email',
    'leader_email': 'leader_email',
    'leader phone': 'leader_phone',
    'leader_phone': 'leader_phone',
    'team size': 'team_size',
    'team_size': 'team_size',
    'problem statements': 'problem_statement',
    'problem statement': 'problem_statement',
    'problem_statement': 'problem_statement',
    'problem_statements': 'problem_statement',
    'upload supporting screenshot of your selected track': 'upload_screenshot',
    'share the link for your skill badge': 'upload_screenshot',
    'upload_screenshot': 'upload_screenshot',
    'screenshot': 'upload_screenshot',
    # Cohort 2 Code Lab Action Center tabs use "Upload File" (GCS screenshot URL)
    'upload file': 'upload_screenshot',
    'upload_file': 'upload_screenshot',
    'file link': 'upload_screenshot',
    'file url': 'upload_screenshot',
    'upload': 'upload_screenshot',
    'created at': 'created_at',
    'created_at': 'created_at',
    'created by name': 'created_by_name',
    'created_by_name': 'created_by_name',
    'created by email': 'created_by_email',
    'created_by_email': 'created_by_email',
    'updated at': 'updated_at',
    'updated_at': 'updated_at',
    'updated by name': 'updated_by_name',
    'updated_by_name': 'updated_by_name',
    'updated by email': 'updated_by_email',
    'updated_by_email': 'updated_by_email',
}


def _map_submission_columns(columns):
    """
    Map XLSX columns to SkillLabSubmission model fields.
    Returns dict: { xlsx_column_name: model_field_name }.
    """
    mapping = {}
    for col in columns:
        if col is None:
            continue
        key = str(col).strip().lower()
        if key in _SUBMISSION_COL_MAP:
            mapping[col] = _SUBMISSION_COL_MAP[key]
    return mapping


def _find_leader_email_column(columns):
    """Return the XLSX column name that maps to leader_email, or None."""
    for col in columns:
        if col is None:
            continue
        key = str(col).strip().lower()
        if key in ('leader email', 'leader_email'):
            return col
    # Fallback: any column containing 'leader' and 'email'
    for col in columns:
        if col is None:
            continue
        low = str(col).strip().lower()
        if 'leader' in low and 'email' in low:
            return col
    return None


def _find_leader_name_column(columns):
    """Return the XLSX column name that maps to leader_name, or None."""
    for col in columns:
        if col is None:
            continue
        key = str(col).strip().lower()
        if key in ('leader name', 'leader_name'):
            return col
    for col in columns:
        if col is None:
            continue
        low = str(col).strip().lower()
        if 'leader' in low and 'name' in low and 'email' not in low and 'phone' not in low:
            return col
    return None


def _find_leader_phone_column(columns):
    """Return the XLSX column name that maps to leader_phone, or None."""
    for col in columns:
        if col is None:
            continue
        key = str(col).strip().lower()
        if key in ('leader phone', 'leader_phone'):
            return col
    for col in columns:
        if col is None:
            continue
        low = str(col).strip().lower()
        if 'leader' in low and ('phone' in low or 'mobile' in low):
            return col
    return None


# Project submission sheets: headers match action-center export (leader email -> user_pii.email)
_PROJECT_SUBMISSION_COL_MAP = {
    'team name': 'team_name',
    'team_name': 'team_name',
    'leader name': 'leader_name',
    'leader_name': 'leader_name',
    'leader email': 'leader_email',
    'leader_email': 'leader_email',
    'leader phone': 'leader_phone',
    'leader_phone': 'leader_phone',
    'team size': 'team_size',
    'team_size': 'team_size',
    'problem statements': 'problem_statement',
    'problem statement': 'problem_statement',
    'problem_statement': 'problem_statement',
    'problem_statements': 'problem_statement',
    'cloud run deployment link': 'cloud_run_deployment_link',
    'cloud_run_deployment_link': 'cloud_run_deployment_link',
    'github repository link': 'github_repository_link',
    'github_repository_link': 'github_repository_link',
    'demo video link': 'demo_video_link',
    'demo_video_link': 'demo_video_link',
    'final project ppt': 'final_project_ppt',
    'final_project_ppt': 'final_project_ppt',
    'created at': 'created_at',
    'created_at': 'created_at',
    'created by name': 'created_by_name',
    'created_by_name': 'created_by_name',
    'created by email': 'created_by_email',
    'created_by_email': 'created_by_email',
    'updated at': 'updated_at',
    'updated_at': 'updated_at',
    'updated by name': 'updated_by_name',
    'updated_by_name': 'updated_by_name',
    'updated by email': 'updated_by_email',
    'updated_by_email': 'updated_by_email',
    'score': 'score',
    'project score': 'score',
    'project_score': 'score',
    'final score': 'score',
    'final_score': 'score',
    'total score': 'score',
    'total_score': 'score',
}


def _find_score_column(columns):
    """Return the column that holds numeric project score, or None."""
    preferred = (
        'score',
        'project score',
        'final score',
        'total score',
        'project_score',
        'final_score',
        'total_score',
    )
    lowered = {}
    for col in columns:
        if col is None:
            continue
        k = str(col).strip().lower()
        lowered[k] = col
    for p in preferred:
        if p in lowered:
            return lowered[p]
    for col in columns:
        if col is None:
            continue
        low = str(col).strip().lower()
        if 'score' in low and 'email' not in low:
            return col
    return None


def _parse_project_score_value(val):
    """Parse a cell value into a float suitable for project_submission.score, or None."""
    if val is None:
        return None
    try:
        if pd.isna(val):
            return None
    except Exception:
        pass
    if isinstance(val, bool):
        return None
    if isinstance(val, (int, float)):
        return float(val)
    s = str(val).strip()
    if not s:
        return None
    s = s.rstrip('%').strip().replace(',', '')
    try:
        return float(s)
    except (ValueError, TypeError):
        return None


def _map_project_submission_columns(columns):
    """Map XLSX columns to ProjectSubmission model fields (excludes track_number, set from sheet)."""
    mapping = {}
    for col in columns:
        if col is None:
            continue
        key = str(col).strip().lower()
        if key in _PROJECT_SUBMISSION_COL_MAP:
            mapping[col] = _PROJECT_SUBMISSION_COL_MAP[key]
    return mapping


def import_project_submission(df, track_number, progress_callback=None):
    """
    Import one Project Submission track sheet into project_submission.
    Upsert by (leader_email, track_number). Leader email must exist in user_pii.
    """
    from sqlalchemy import func
    from server.models import db

    IM = _import_models()
    UserPII = IM.UserPII
    ProjectSubmission = IM.ProjectSubmission

    columns = list(df.columns)
    col_map = _map_project_submission_columns(columns)
    leader_email_col = _find_leader_email_column(columns)
    leader_name_col = _find_leader_name_column(columns)
    leader_phone_col = _find_leader_phone_column(columns)

    if not leader_email_col:
        raise Exception(
            "Could not find a 'Leader Email' column in the project submission sheet. "
            "Please ensure the sheet has a column named 'Leader Email'."
        )

    total_rows = len(df)
    created = 0
    updated = 0
    skipped = 0
    errors = []
    rows_errors = []

    pii_pre = _ensure_user_pii_for_leaders(
        df, leader_email_col, name_col=leader_name_col, phone_col=leader_phone_col,
    )

    emails_in_sheet = []
    try:
        for v in df[leader_email_col].dropna():
            e = str(v).strip().lower()
            if e and '@' in e:
                emails_in_sheet.append(e)
    except Exception:
        pass
    unique_sheet_emails = list(dict.fromkeys(emails_in_sheet))

    pii_by_lower = {}
    try:
        for chunk in _chunk_list(unique_sheet_emails, IMPORT_EMAIL_IN_CHUNK):
            if not chunk:
                continue
            for p in UserPII.query.filter(func.lower(UserPII.email).in_(chunk)).all():
                el = (p.email or '').strip().lower()
                if el:
                    pii_by_lower[el] = p
    except Exception:
        pass

    existing = {}
    try:
        for chunk in _chunk_list(unique_sheet_emails, IMPORT_EMAIL_IN_CHUNK):
            if not chunk:
                continue
            for ps_row in ProjectSubmission.query.filter(
                ProjectSubmission.track_number == track_number,
                func.lower(ProjectSubmission.leader_email).in_(chunk),
            ).all():
                existing[(ps_row.leader_email or '').lower()] = ps_row
    except Exception:
        pass

    link_fields = (
        'cloud_run_deployment_link', 'github_repository_link', 'demo_video_link', 'final_project_ppt',
    )
    short_str_fields = (
        'team_name', 'leader_name', 'leader_phone', 'created_by_name', 'created_by_email',
        'updated_by_name', 'updated_by_email',
    )

    for index, row in df.iterrows():
        email_val = row.get(leader_email_col)
        if pd.isna(email_val):
            skipped += 1
            _record_row_error(errors, rows_errors, _classify_row_error(
                index, None, raw_email='', reason_code='missing_email',
                reason_message='Missing leader email',
            ))
            continue
        email_norm = str(email_val).strip().lower()
        if not email_norm or '@' not in email_norm:
            skipped += 1
            _record_row_error(errors, rows_errors, _classify_row_error(
                index, None, raw_email=email_norm, reason_code='invalid_email',
                reason_message='Invalid leader email',
            ))
            continue

        pii = pii_by_lower.get(email_norm)
        if not pii:
            skipped += 1
            _record_row_error(errors, rows_errors, _classify_row_error(
                index, None, raw_email=email_norm, reason_code='fk_user_pii_missing',
                reason_message=f'Leader email not found in user PII ({email_norm})',
            ))
            continue

        canonical_email = (pii.email or '').strip()
        if not canonical_email:
            skipped += 1
            _record_row_error(errors, rows_errors, _classify_row_error(
                index, None, raw_email=email_norm, reason_code='fk_user_pii_missing',
                reason_message='Empty user PII email',
            ))
            continue

        try:
            with db.session.begin_nested():
                data = {'track_number': track_number}
                for xlsx_col, model_field in col_map.items():
                    if model_field == 'leader_email':
                        continue
                    val = row.get(xlsx_col)
                    if pd.isna(val):
                        val = None
                    elif model_field == 'team_size':
                        try:
                            val = int(val)
                        except (ValueError, TypeError):
                            val = None
                    elif model_field == 'score':
                        val = _parse_project_score_value(val)
                    elif model_field in ('created_at', 'updated_at'):
                        if val is not None:
                            try:
                                if isinstance(val, str):
                                    val = pd.to_datetime(val)
                                elif not isinstance(val, datetime):
                                    val = pd.to_datetime(val)
                            except Exception:
                                val = None
                    else:
                        val = str(val).strip() if val is not None else None
                        if val and model_field in short_str_fields:
                            val = val[:255]
                        elif val and model_field in link_fields:
                            val = val[:1024]
                        elif val and model_field == 'problem_statement':
                            pass
                    data[model_field] = val

                data['leader_email'] = canonical_email

                rec = existing.get(canonical_email.lower())
                if rec:
                    for field, val in data.items():
                        if field in ('leader_email', 'track_number'):
                            continue
                        if val is not None:
                            setattr(rec, field, val)
                    rec.updated_at = datetime.utcnow()
                    updated += 1
                else:
                    rec = ProjectSubmission(**data)
                    db.session.add(rec)
                    existing[canonical_email.lower()] = rec
                    created += 1

            if progress_callback:
                try:
                    progress_callback(created, updated, skipped)
                except Exception:
                    pass

        except Exception as e:
            skipped += 1
            _record_row_error(errors, rows_errors, _classify_row_error(
                index, e, raw_email=email_norm,
            ))

    try:
        db.session.commit()
    except Exception:
        db.session.rollback()

    return {
        'total_rows': total_rows,
        'created': created,
        'updated': updated,
        'skipped': skipped,
        'errors': errors[:100],
        'rows_errors': rows_errors,
        'pii_auto_created': pii_pre.get('pii_auto_created', 0),
        'pii_auto_skipped': pii_pre.get('pii_auto_skipped', 0),
    }


def import_project_submission_scores(df, track_number, progress_callback=None, audit_user=None):
    """
    Set project_submission.score from a CSV/Excel with leader email + score columns.
    Only updates rows that already exist for (leader_email, track_number). Leader email
    must match a user_pii row (canonical casing applied).

    audit_user: optional dict with 'name' and 'email' for updated_by_* on each row.
    """
    from sqlalchemy import func
    from server.models import db

    if track_number not in (1, 2, 3):
        raise ValueError("track_number must be 1, 2, or 3")

    IM = _import_models()
    UserPII = IM.UserPII
    ProjectSubmission = IM.ProjectSubmission

    columns = list(df.columns)
    leader_email_col = _find_leader_email_column(columns)
    score_col = _find_score_column(columns)

    if not leader_email_col:
        raise Exception(
            "Could not find a leader email column. Use a header such as 'Leader Email' or 'leader_email'."
        )
    if not score_col:
        raise Exception(
            "Could not find a score column. Use a header such as 'Score' or 'Project Score'."
        )

    total_rows = len(df)
    updated = 0
    skipped = 0
    errors = []

    emails_in_sheet = []
    try:
        for v in df[leader_email_col].dropna():
            e = str(v).strip().lower()
            if e and '@' in e:
                emails_in_sheet.append(e)
    except Exception:
        pass
    unique_sheet_emails = list(dict.fromkeys(emails_in_sheet))

    pii_by_lower = {}
    try:
        for chunk in _chunk_list(unique_sheet_emails, IMPORT_EMAIL_IN_CHUNK):
            if not chunk:
                continue
            for p in UserPII.query.filter(func.lower(UserPII.email).in_(chunk)).all():
                el = (p.email or '').strip().lower()
                if el:
                    pii_by_lower[el] = p
    except Exception:
        pass

    existing = {}
    try:
        for chunk in _chunk_list(unique_sheet_emails, IMPORT_EMAIL_IN_CHUNK):
            if not chunk:
                continue
            for ps_row in ProjectSubmission.query.filter(
                ProjectSubmission.track_number == track_number,
                func.lower(ProjectSubmission.leader_email).in_(chunk),
            ).all():
                existing[(ps_row.leader_email or '').lower()] = ps_row
    except Exception:
        pass

    for index, row in df.iterrows():
        try:
            email_val = row.get(leader_email_col)
            if pd.isna(email_val):
                skipped += 1
                if len(errors) < 100:
                    errors.append(f"Row {index + 2}: Missing leader email")
                continue
            email_norm = str(email_val).strip().lower()
            if not email_norm or '@' not in email_norm:
                skipped += 1
                if len(errors) < 100:
                    errors.append(f"Row {index + 2}: Invalid leader email")
                continue

            pii = pii_by_lower.get(email_norm)
            if not pii:
                skipped += 1
                if len(errors) < 100:
                    errors.append(f"Row {index + 2}: Leader email not in user PII ({email_norm})")
                continue

            canonical_email = (pii.email or '').strip()
            if not canonical_email:
                skipped += 1
                if len(errors) < 100:
                    errors.append(f"Row {index + 2}: Empty user PII email")
                continue

            score_val = _parse_project_score_value(row.get(score_col))
            if score_val is None:
                skipped += 1
                if len(errors) < 100:
                    errors.append(f"Row {index + 2}: Missing or invalid score")
                continue

            rec = existing.get(canonical_email.lower())
            if not rec:
                skipped += 1
                if len(errors) < 100:
                    errors.append(
                        f"Row {index + 2}: No project submission for track {track_number} ({email_norm})"
                    )
                continue

            rec.score = score_val
            rec.updated_at = datetime.utcnow()
            if audit_user:
                rec.updated_by_name = (audit_user.get('name') or '')[:255] or None
                rec.updated_by_email = (audit_user.get('email') or '')[:255] or None
            updated += 1

            if updated % IMPORT_BATCH_COMMIT_SIZE == 0:
                db.session.commit()

            if progress_callback:
                try:
                    progress_callback(updated, skipped)
                except Exception:
                    pass

        except Exception as e:
            db.session.rollback()
            skipped += 1
            if len(errors) < 100:
                errors.append(f"Row {index + 2}: {str(e)}")

    try:
        db.session.commit()
    except Exception:
        db.session.rollback()

    return {
        'total_rows': total_rows,
        'updated': updated,
        'skipped': skipped,
        'errors': errors[:100],
    }


def _skilllab_submission_fingerprint(email: str, data: dict) -> tuple:
    """Unique import key: leader email + submission link (upload_screenshot)."""
    link = (data.get('upload_screenshot') or '').strip().lower()
    return (email, link)


def _load_skilllab_import_fingerprints(SkillLabSubmission, emails_in_sheet):
    """Fingerprints of rows already stored for leaders in this import."""
    from sqlalchemy import func

    fingerprints = set()
    if not emails_in_sheet:
        return fingerprints
    try:
        for chunk in _chunk_list(emails_in_sheet, IMPORT_EMAIL_IN_CHUNK):
            if not chunk:
                continue
            for srow in SkillLabSubmission.query.filter(
                func.lower(SkillLabSubmission.leader_email).in_(chunk)
            ).all():
                fp = _skilllab_submission_fingerprint(
                    (srow.leader_email or '').strip().lower(),
                    {'upload_screenshot': srow.upload_screenshot},
                )
                fingerprints.add(fp)
    except Exception:
        pass
    return fingerprints


def import_skilllab_submission(df, progress_callback=None):
    """
    Import Skill Lab submissions from a DataFrame into skilllab_submission.
    Inserts one row per sheet row; multiple rows per leader_email are kept when
    the submission link differs. Re-import skips rows that already exist for the
    same (leader_email, upload_screenshot) pair.

    Verification fields (valid, remark) on existing rows are never overwritten.
    Metrics use only the latest valid submission per leader and track; see
    server.utils.skilllab_submission_selection.

    Robustness: every row runs in its own SAVEPOINT
    (db.session.begin_nested) so a single FK / integrity error never poisons
    the rest of the batch. Missing user_pii rows for leaders in the sheet are
    auto-created up front.

    Returns dict with total_rows, created, updated, skipped, skipped_duplicate,
    errors, rows_errors, pii_auto_created, pii_auto_skipped.
    """
    from server.models import db

    IM = _import_models()
    SkillLabSubmission = IM.SkillLabSubmission

    columns = list(df.columns)
    col_map = _map_submission_columns(columns)
    leader_email_col = _find_leader_email_column(columns)
    leader_name_col = _find_leader_name_column(columns)
    leader_phone_col = _find_leader_phone_column(columns)

    if not leader_email_col:
        raise Exception(
            "Could not find a 'Leader Email' column in the submission sheet. "
            "Please ensure the sheet has a column named 'Leader Email'."
        )

    total_rows = len(df)
    created = 0
    updated = 0
    skipped = 0
    skipped_duplicate = 0
    errors = []
    rows_errors = []

    pii_pre = _ensure_user_pii_for_leaders(
        df, leader_email_col, name_col=leader_name_col, phone_col=leader_phone_col,
    )

    emails_in_sheet = set()
    try:
        for v in df[leader_email_col].dropna():
            e = str(v).strip().lower()
            if e and '@' in e:
                emails_in_sheet.add(e)
    except Exception:
        pass

    existing_fingerprints = _load_skilllab_import_fingerprints(
        SkillLabSubmission, emails_in_sheet,
    )

    for index, row in df.iterrows():
        email_val = row.get(leader_email_col)
        if pd.isna(email_val):
            skipped += 1
            _record_row_error(errors, rows_errors, _classify_row_error(
                index, None, raw_email='', reason_code='missing_email',
                reason_message='Missing leader email',
            ))
            continue
        email = str(email_val).strip().lower()
        if not email or '@' not in email:
            skipped += 1
            _record_row_error(errors, rows_errors, _classify_row_error(
                index, None, raw_email=email, reason_code='invalid_email',
                reason_message='Invalid leader email',
            ))
            continue

        try:
            with db.session.begin_nested():
                data = {}
                for xlsx_col, model_field in col_map.items():
                    val = row.get(xlsx_col)
                    if pd.isna(val):
                        val = None
                    elif model_field in ('team_size',):
                        try:
                            val = int(val)
                        except (ValueError, TypeError):
                            val = None
                    elif model_field in ('created_at', 'updated_at'):
                        if val is not None:
                            try:
                                if isinstance(val, str):
                                    val = pd.to_datetime(val)
                                elif not isinstance(val, datetime):
                                    val = pd.to_datetime(val)
                            except Exception:
                                val = None
                    else:
                        val = str(val).strip() if val is not None else None
                        if val and model_field in ('team_name', 'leader_name', 'leader_phone',
                                                    'created_by_name', 'created_by_email',
                                                    'updated_by_name', 'updated_by_email'):
                            val = val[:255]
                        elif val and model_field == 'upload_screenshot':
                            val = val[:1024]
                    data[model_field] = val

                data['leader_email'] = email

                fp = _skilllab_submission_fingerprint(email, data)
                if fp in existing_fingerprints:
                    skipped_duplicate += 1
                    continue

                rec = SkillLabSubmission(**data)
                db.session.add(rec)
                existing_fingerprints.add(fp)
                created += 1

            if progress_callback:
                try:
                    progress_callback(created, updated, skipped)
                except Exception:
                    pass

        except Exception as e:
            skipped += 1
            _record_row_error(errors, rows_errors, _classify_row_error(
                index, e, raw_email=email,
            ))

    try:
        db.session.commit()
    except Exception:
        db.session.rollback()

    return {
        'total_rows': total_rows,
        'created': created,
        'updated': updated,
        'skipped': skipped,
        'skipped_duplicate': skipped_duplicate,
        'errors': errors[:100],
        'rows_errors': rows_errors,
        'pii_auto_created': pii_pre.get('pii_auto_created', 0),
        'pii_auto_skipped': pii_pre.get('pii_auto_skipped', 0),
    }


def import_codelab_submission(
    df,
    progress_callback=None,
    sheet_track_number=None,
    default_problem_statement=None,
):
    """
    Import Code Lab submissions from a DataFrame into codelab_submission.

    Cohort 1 (single "Code Lab Submission" tab): upsert by leader_email only.
    Cohort 2 (per-track tabs): upsert by (leader_email, track_number,
    problem_statement) when sheet_track_number and/or default_problem_statement
    are set (Professional Track N / Student Track sheets).

    Per-row SAVEPOINT (db.session.begin_nested) isolates row failures and
    missing user_pii rows for leaders are auto-created up front.

    Returns dict with total_rows, created, updated, skipped, errors,
    rows_errors, pii_auto_created, pii_auto_skipped.
    """
    from server.models import db

    IM = _import_models()
    CodeLabSubmission = IM.CodeLabSubmission

    columns = list(df.columns)
    col_map = _map_submission_columns(columns)
    leader_email_col = _find_leader_email_column(columns)
    leader_name_col = _find_leader_name_column(columns)
    leader_phone_col = _find_leader_phone_column(columns)

    if not leader_email_col:
        raise Exception(
            "Could not find a 'Leader Email' column in the Code Lab submission sheet. "
            "Please ensure the sheet has a column named 'Leader Email'."
        )

    total_rows = len(df)
    created = 0
    updated = 0
    skipped = 0
    errors = []
    rows_errors = []

    from sqlalchemy import func

    pii_pre = _ensure_user_pii_for_leaders(
        df, leader_email_col, name_col=leader_name_col, phone_col=leader_phone_col,
    )

    emails_in_sheet = set()
    try:
        for v in df[leader_email_col].dropna():
            e = str(v).strip().lower()
            if e and '@' in e:
                emails_in_sheet.add(e)
    except Exception:
        pass

    # Cohort 2 per-track tabs: one row per (email, track). Cohort 1: one row per email.
    # Key by track (not problem_statement) so re-imports still match rows that were
    # stored with a null/empty lab label before defaults were filled in.
    use_track_key = sheet_track_number is not None
    use_ps_key = (not use_track_key) and default_problem_statement is not None
    existing = {}
    try:
        for chunk in _chunk_list(emails_in_sheet, IMPORT_EMAIL_IN_CHUNK):
            if not chunk:
                continue
            for crow in CodeLabSubmission.query.filter(
                func.lower(CodeLabSubmission.leader_email).in_(chunk)
            ).all():
                email_key = crow.leader_email.lower()
                if use_track_key:
                    existing[(email_key, crow.track_number)] = crow
                elif use_ps_key:
                    existing[(email_key, crow.problem_statement or '')] = crow
                else:
                    existing[email_key] = crow
    except Exception:
        pass

    for index, row in df.iterrows():
        email_val = row.get(leader_email_col)
        if pd.isna(email_val):
            skipped += 1
            _record_row_error(errors, rows_errors, _classify_row_error(
                index, None, raw_email='', reason_code='missing_email',
                reason_message='Missing leader email',
            ))
            continue
        email = str(email_val).strip().lower()
        if not email or '@' not in email:
            skipped += 1
            _record_row_error(errors, rows_errors, _classify_row_error(
                index, None, raw_email=email, reason_code='invalid_email',
                reason_message='Invalid leader email',
            ))
            continue

        try:
            with db.session.begin_nested():
                data = {}
                for xlsx_col, model_field in col_map.items():
                    val = row.get(xlsx_col)
                    if pd.isna(val):
                        val = None
                    elif model_field in ('team_size',):
                        try:
                            val = int(val)
                        except (ValueError, TypeError):
                            val = None
                    elif model_field in ('created_at', 'updated_at'):
                        if val is not None:
                            try:
                                if isinstance(val, str):
                                    val = pd.to_datetime(val)
                                elif not isinstance(val, datetime):
                                    val = pd.to_datetime(val)
                            except Exception:
                                val = None
                    else:
                        val = str(val).strip() if val is not None else None
                        if val and model_field in ('team_name', 'leader_name', 'leader_phone',
                                                    'created_by_name', 'created_by_email',
                                                    'updated_by_name', 'updated_by_email'):
                            val = val[:255]
                        elif val and model_field == 'upload_screenshot':
                            val = val[:1024]
                    data[model_field] = val

                data['leader_email'] = email
                if sheet_track_number is not None:
                    data['track_number'] = sheet_track_number
                if default_problem_statement and not data.get('problem_statement'):
                    data['problem_statement'] = default_problem_statement

                if use_track_key:
                    lookup_key = (email, data.get('track_number'))
                    # Legacy student rows were stored with track_number NULL.
                    rec = existing.get(lookup_key)
                    if rec is None and sheet_track_number == 3:
                        rec = existing.get((email, None))
                elif use_ps_key:
                    lookup_key = (email, data.get('problem_statement') or '')
                    rec = existing.get(lookup_key)
                else:
                    lookup_key = email
                    rec = existing.get(email)
                if rec:
                    for field, val in data.items():
                        if field == 'leader_email':
                            continue
                        if val is not None:
                            setattr(rec, field, val)
                    rec.updated_at = datetime.utcnow()
                    updated += 1
                    if use_track_key:
                        existing[(email, data.get('track_number'))] = rec
                else:
                    rec = CodeLabSubmission(**data)
                    db.session.add(rec)
                    existing[lookup_key] = rec
                    created += 1

            if progress_callback:
                try:
                    progress_callback(created, updated, skipped)
                except Exception:
                    pass

        except Exception as e:
            skipped += 1
            _record_row_error(errors, rows_errors, _classify_row_error(
                index, e, raw_email=email,
            ))

    try:
        db.session.commit()
    except Exception:
        db.session.rollback()

    return {
        'total_rows': total_rows,
        'created': created,
        'updated': updated,
        'skipped': skipped,
        'errors': errors[:100],
        'rows_errors': rows_errors,
        'pii_auto_created': pii_pre.get('pii_auto_created', 0),
        'pii_auto_skipped': pii_pre.get('pii_auto_skipped', 0),
    }


def import_codelab_submission_sheets(codelab_sheets, progress_callback=None):
    """Import one or more Code Lab worksheets; aggregate counts for the API response."""
    totals = {
        'total_rows': 0,
        'created': 0,
        'updated': 0,
        'skipped': 0,
        'errors': [],
        'rows_errors': [],
        'pii_auto_created': 0,
        'pii_auto_skipped': 0,
        'sheets': [],
    }
    if not codelab_sheets:
        return totals

    for sheet_info in codelab_sheets:
        df = sheet_info.get('df')
        if not _nonempty_dataframe(df):
            continue
        one = import_codelab_submission(
            df,
            progress_callback=progress_callback,
            sheet_track_number=sheet_info.get('track'),
            default_problem_statement=sheet_info.get('default_problem_statement'),
        )
        one['sheet_name'] = sheet_info.get('sheet_name', '')
        one['track'] = sheet_info.get('track')
        totals['sheets'].append(one)
        totals['total_rows'] += one.get('total_rows', 0) or 0
        totals['created'] += one.get('created', 0) or 0
        totals['updated'] += one.get('updated', 0) or 0
        totals['skipped'] += one.get('skipped', 0) or 0
        totals['pii_auto_created'] += one.get('pii_auto_created', 0) or 0
        totals['pii_auto_skipped'] += one.get('pii_auto_skipped', 0) or 0
        if one.get('errors'):
            totals['errors'].extend(one['errors'])
        if one.get('rows_errors'):
            totals['rows_errors'].extend(one['rows_errors'])
    totals['errors'] = totals['errors'][:100]
    return totals


def import_lab_completion_sheet(df, track_number, lab_number, progress_callback=None):
    """
    Import a Lab Completion sheet into codelab_submission.
    Each row needs an email column. Upsert by (leader_email, track_number, problem_statement).
    problem_statement is set to "Lab 1" or "Lab 2" based on lab_number.
    """
    from server.models import db

    IM = _import_models()
    CodeLabSubmission = IM.CodeLabSubmission

    columns = list(df.columns)
    email_col = _find_email_column(columns)
    if not email_col:
        raise Exception(
            "Could not find an Email column in the Lab Completion sheet. "
            "Please ensure the sheet has a column named 'Email'."
        )

    # Try to find a name column
    name_col = None
    for col in columns:
        if col is None:
            continue
        low = str(col).strip().lower()
        if low in ('name', 'full name', 'fullname', 'leader name', 'leader_name'):
            name_col = col
            break
    if not name_col:
        for col in columns:
            if col and 'name' in str(col).lower() and 'email' not in str(col).lower():
                name_col = col
                break

    # Try to find an upload/file/screenshot column
    upload_col = None
    for col in columns:
        if col is None:
            continue
        low = str(col).strip().lower()
        if low in ('upload file', 'upload_file', 'upload', 'file', 'screenshot',
                    'upload_screenshot', 'upload screenshot', 'file link', 'file url'):
            upload_col = col
            break
    if not upload_col:
        for col in columns:
            if col and ('upload' in str(col).lower() or 'screenshot' in str(col).lower()):
                upload_col = col
                break

    problem_stmt = f"Lab {lab_number}"

    total_rows = len(df)
    created = 0
    updated = 0
    skipped = 0
    errors = []
    rows_errors = []

    pii_pre = _ensure_user_pii_for_leaders(df, email_col, name_col=name_col)

    emails_in_sheet = set()
    try:
        for v in df[email_col].dropna():
            e = str(v).strip().lower()
            if e and '@' in e:
                emails_in_sheet.add(e)
    except Exception:
        pass

    existing = {}
    try:
        for chunk in _chunk_list(emails_in_sheet, IMPORT_EMAIL_IN_CHUNK):
            if not chunk:
                continue
            for row in CodeLabSubmission.query.filter(
                CodeLabSubmission.track_number == track_number,
                CodeLabSubmission.problem_statement == problem_stmt,
                CodeLabSubmission.leader_email.in_(chunk),
            ).all():
                existing[row.leader_email.lower()] = row
    except Exception:
        pass

    for index, row in df.iterrows():
        email_val = row.get(email_col)
        if pd.isna(email_val):
            skipped += 1
            _record_row_error(errors, rows_errors, _classify_row_error(
                index, None, raw_email='', reason_code='missing_email',
                reason_message='Missing email',
            ))
            continue
        email = str(email_val).strip().lower()
        if not email or '@' not in email:
            skipped += 1
            _record_row_error(errors, rows_errors, _classify_row_error(
                index, None, raw_email=email, reason_code='invalid_email',
                reason_message='Invalid email',
            ))
            continue

        try:
            with db.session.begin_nested():
                name_val = None
                if name_col:
                    nv = row.get(name_col)
                    if not pd.isna(nv):
                        name_val = str(nv).strip()[:255] or None

                upload_val = None
                if upload_col:
                    uv = row.get(upload_col)
                    if not pd.isna(uv):
                        upload_val = str(uv).strip()[:1024] or None

                rec = existing.get(email)
                if rec:
                    if name_val and not rec.leader_name:
                        rec.leader_name = name_val
                    if upload_val:
                        rec.upload_screenshot = upload_val
                    rec.updated_at = datetime.utcnow()
                    updated += 1
                else:
                    rec = CodeLabSubmission(
                        leader_email=email,
                        leader_name=name_val,
                        track_number=track_number,
                        problem_statement=problem_stmt,
                        upload_screenshot=upload_val,
                    )
                    db.session.add(rec)
                    existing[email] = rec
                    created += 1

            if progress_callback:
                try:
                    progress_callback(created, updated, skipped)
                except Exception:
                    pass

        except Exception as e:
            skipped += 1
            _record_row_error(errors, rows_errors, _classify_row_error(
                index, e, raw_email=email,
            ))

    try:
        db.session.commit()
    except Exception:
        db.session.rollback()

    return {
        'total_rows': total_rows,
        'created': created,
        'updated': updated,
        'skipped': skipped,
        'errors': errors[:100],
        'rows_errors': rows_errors,
        'pii_auto_created': pii_pre.get('pii_auto_created', 0),
        'pii_auto_skipped': pii_pre.get('pii_auto_skipped', 0),
    }


def _find_mcq_question_columns(columns, email_col):
    """
    Resolve 10 question columns from Excel headers.
    Returns list of (xlsx_column_name, model_field) for question_1..question_10.
    Matches headers that start with "Q1.", "Q2.", ... "Q10." (Excel often has "Q1. What", "Q2. Why...").
    Or exact "q1", "question 1", etc.
    Fallback: columns G–P (7th–16th column) are the 10 answer columns for all tracks.
    Last fallback: first 10 non-email columns.
    """
    col_list = [c for c in columns if c is not None and str(c).strip()]
    email_col_str = email_col.strip().lower() if email_col else ''
    model_fields = ['question_1', 'question_2', 'question_3', 'question_4', 'question_5',
                    'question_6', 'question_7', 'question_8', 'question_9', 'question_10']

    # Match by prefix: "Q1. What" -> question_1, "Q10. Why" -> question_10. Check Q10 before Q1.
    result = [None] * 10
    for col in col_list:
        key = str(col).strip().lower()
        if key == email_col_str:
            continue
        key_ns = key.replace(' ', '')
        for i in range(10, 0, -1):  # 10 down to 1 so "q10." matches before "q1."
            prefix = f'q{i}.'
            if key.startswith(prefix) or key_ns.startswith(prefix):
                if result[i - 1] is None:
                    result[i - 1] = (col, model_fields[i - 1])
                break

    out = [x for x in result if x is not None]
    if len(out) >= 10:
        return out[:10]

    # Exact matches for any remaining slots (e.g. "q1", "question 1")
    norm_to_col = {str(c).strip().lower(): c for c in col_list if str(c).strip().lower() != email_col_str}
    norm_to_col_no_space = {k.replace(' ', ''): v for k, v in norm_to_col.items()}
    for i in range(1, 11):
        if result[i - 1] is not None:
            continue
        mf = model_fields[i - 1]
        for cand in [f'q{i}.', f'q{i}', f'question {i}', str(i), f'answer {i}', f'question{i}', f'q {i}', f'answer{i}']:
            if cand in norm_to_col:
                result[i - 1] = (norm_to_col[cand], mf)
                break
            if cand.replace(' ', '') in norm_to_col_no_space:
                result[i - 1] = (norm_to_col_no_space[cand.replace(' ', '')], mf)
                break

    out = [x for x in result if x is not None]
    if len(out) >= 10:
        return out[:10]

    # Fallback: columns G through P (7th–16th column, 0-based indices 6–15) are the 10 answer columns for all tracks
    cols = list(columns)
    if len(cols) >= 16:
        return [(cols[6 + i], model_fields[i]) for i in range(10)]

    # Fallback: positional (first 10 non-email columns)
    result = []
    for c in col_list:
        if str(c).strip().lower() == email_col_str:
            continue
        if len(result) >= 10:
            break
        result.append((c, model_fields[len(result)]))
    return result[:10]


def _find_mcq_column(columns, *candidates):
    """Return first column whose name (normalized, lower) matches one of the candidates."""
    for col in columns:
        if col is None:
            continue
        key = ' '.join(str(col).strip().lower().split())
        for c in candidates:
            cnorm = ' '.join(c.strip().lower().split())
            if cnorm == key or key == cnorm or key.endswith(cnorm) or cnorm in key:
                return col
    return None


def _parse_mcq_score_slash(cell_value):
    """
    Parse Action Center score cells like '6 / 10' or '0/10'.
    Returns (numerator, denominator) integers, or (None, None) if not parseable.
    """
    if cell_value is None or (isinstance(cell_value, float) and pd.isna(cell_value)):
        return None, None
    s = str(cell_value).strip()
    m = re.match(r"^(\d+)\s*/\s*(\d+)", s)
    if m:
        return int(m.group(1)), int(m.group(2))
    try:
        v = int(float(s))
        return v, 10
    except (ValueError, TypeError):
        return None, None


def _find_question_columns_score_to_created(columns):
    """
    For Cohort 2 optional sheet: ten question columns sit between 'Score' and 'Created At'.
    Returns list of (xlsx_col, question_N) or None.
    """
    cols = list(columns)
    score_col = _find_mcq_column(columns, "score")
    created_col = _find_mcq_column(columns, "created at", "Created At")
    if not score_col or not created_col:
        return None
    try:
        i0 = cols.index(score_col)
        i1 = cols.index(created_col)
    except ValueError:
        return None
    if i1 <= i0 + 1:
        return None
    middle = cols[i0 + 1:i1]
    if len(middle) < 10:
        return None
    model_fields = [f"question_{i}" for i in range(1, 11)]
    return [(middle[i], model_fields[i]) for i in range(10)]


def import_optional_mcq_response(
    df,
    track_number,
    progress_callback=None,
    score_from_sheet=False,
    allow_multiple_per_email=False,
):
    """
    Import Optional MCQ responses from a DataFrame into optional_mcq_response.
    Upsert by (track_number, email): update if exists, else insert — unless allow_multiple_per_email
    is True (Cohort 2 track 4), in which case every row is inserted and duplicates per email are kept.
    Returns dict: total_rows, created, updated, skipped, errors.
    """
    from server.models import db

    IM = _import_models()
    OptionalMcqResponse = IM.OptionalMcqResponse

    columns = list(df.columns)
    email_col = _find_leader_email_column(columns)
    if not email_col:
        raise Exception(
            "Could not find a 'Leader Email' column in the MCQ sheet. "
            "Please ensure the sheet has a column named 'Leader Email'."
        )
    question_cols = _find_mcq_question_columns(columns, email_col)
    if not question_cols and score_from_sheet:
        question_cols = _find_question_columns_score_to_created(columns)
    if not question_cols:
        raise Exception("Could not find 10 question columns in the MCQ sheet.")

    score_col = _find_mcq_column(columns, 'score')
    team_name_col = _find_mcq_column(columns, 'team name', 'Team Name')
    leader_name_col = _find_mcq_column(columns, 'leader name', 'Leader Name')
    leader_phone_col = _find_mcq_column(columns, 'leader phone', 'Leader Phone')
    team_size_col = _find_mcq_column(columns, 'team size', 'Team size', 'team_size')
    problem_col = _find_mcq_column(columns, 'problem statements', 'Problem Statements', 'problem statement')
    created_at_col = _find_mcq_column(columns, 'created at', 'Created At')
    created_by_name_col = _find_mcq_column(columns, 'created by name', 'Created By Name')
    created_by_email_col = _find_mcq_column(columns, 'created by email', 'Created By Email')
    updated_at_col = _find_mcq_column(columns, 'updated at', 'Updated At')
    updated_by_name_col = _find_mcq_column(columns, 'updated by name', 'Updated By Name')
    updated_by_email_col = _find_mcq_column(columns, 'updated by email', 'Updated By Email')

    total_rows = len(df)
    created = 0
    updated = 0
    skipped = 0
    errors = []
    rows_errors = []

    from sqlalchemy import func

    pii_pre = _ensure_user_pii_for_leaders(
        df, email_col, name_col=leader_name_col, phone_col=leader_phone_col,
    )

    emails_in_sheet = set()
    existing = {}
    if not allow_multiple_per_email:
        try:
            for v in df[email_col].dropna():
                e = str(v).strip().lower()
                if e and '@' in e:
                    emails_in_sheet.add(e)
        except Exception:
            pass

        try:
            for chunk in _chunk_list(emails_in_sheet, IMPORT_EMAIL_IN_CHUNK):
                if not chunk:
                    continue
                for mrow in OptionalMcqResponse.query.filter(
                    OptionalMcqResponse.track_number == track_number,
                    func.lower(OptionalMcqResponse.email).in_(chunk),
                ).all():
                    existing[mrow.email.lower()] = mrow
        except Exception:
            pass

    for index, row in df.iterrows():
        email_val = row.get(email_col)
        if pd.isna(email_val):
            skipped += 1
            _record_row_error(errors, rows_errors, _classify_row_error(
                index, None, raw_email='', reason_code='missing_email',
                reason_message='Missing email',
            ))
            continue
        email = str(email_val).strip().lower()
        if not email or '@' not in email:
            skipped += 1
            _record_row_error(errors, rows_errors, _classify_row_error(
                index, None, raw_email=email, reason_code='invalid_email',
                reason_message='Invalid email',
            ))
            continue

        try:
            with db.session.begin_nested():
                def _get(col, default=None):
                    if col is None:
                        return default
                    v = row.get(col)
                    if pd.isna(v):
                        return default
                    return str(v).strip() or default

                data = {'track_number': track_number, 'email': email}
                if team_name_col:
                    data['team_name'] = _get(team_name_col)
                if leader_name_col:
                    data['leader_name'] = _get(leader_name_col)
                if leader_phone_col:
                    data['leader_phone'] = _get(leader_phone_col)
                if team_size_col:
                    try:
                        v = row.get(team_size_col)
                        data['team_size'] = int(v) if v is not None and not pd.isna(v) else None
                    except (ValueError, TypeError):
                        data['team_size'] = None
                if problem_col:
                    data['problem_statement'] = _get(problem_col)
                if created_at_col:
                    try:
                        v = row.get(created_at_col)
                        if v is not None and not pd.isna(v):
                            data['created_at'] = pd.to_datetime(v) if hasattr(pd, 'to_datetime') else v
                    except Exception:
                        pass
                if created_by_name_col:
                    data['created_by_name'] = _get(created_by_name_col)
                if created_by_email_col:
                    data['created_by_email'] = _get(created_by_email_col)
                if updated_at_col:
                    try:
                        v = row.get(updated_at_col)
                        if v is not None and not pd.isna(v):
                            data['updated_at'] = pd.to_datetime(v) if hasattr(pd, 'to_datetime') else v
                    except Exception:
                        pass
                if updated_by_name_col:
                    data['updated_by_name'] = _get(updated_by_name_col)
                if updated_by_email_col:
                    data['updated_by_email'] = _get(updated_by_email_col)

                for xlsx_col, model_field in question_cols:
                    val = row.get(xlsx_col)
                    if pd.isna(val):
                        data[model_field] = None
                    else:
                        data[model_field] = str(val).strip() or None

                if score_from_sheet:
                    num, den = _parse_mcq_score_slash(row.get(score_col)) if score_col else (None, None)
                    if num is not None and den and den > 0:
                        data['score'] = min(10, max(0, int(round(num * 10.0 / float(den)))))
                    else:
                        data['score'] = None
                else:
                    from server.utils.mcq_answer_key import score_submission
                    auto = score_submission(
                        track_number,
                        data.get('question_1'), data.get('question_2'), data.get('question_3'), data.get('question_4'),
                        data.get('question_5'), data.get('question_6'), data.get('question_7'), data.get('question_8'),
                        data.get('question_9'), data.get('question_10'),
                    )
                    data['score'] = auto['correct_count']

                rec = None if allow_multiple_per_email else existing.get(email)
                if rec:
                    for k, v in data.items():
                        if k == 'email':
                            continue
                        setattr(rec, k, v)
                    updated += 1
                else:
                    rec = OptionalMcqResponse(**data)
                    db.session.add(rec)
                    if not allow_multiple_per_email:
                        existing[email] = rec
                    created += 1

            if progress_callback:
                try:
                    progress_callback(created, updated, skipped)
                except Exception:
                    pass

        except Exception as e:
            skipped += 1
            _record_row_error(errors, rows_errors, _classify_row_error(
                index, e, raw_email=email,
            ))

    try:
        db.session.commit()
    except Exception:
        db.session.rollback()

    return {
        'total_rows': total_rows,
        'created': created,
        'updated': updated,
        'skipped': skipped,
        'errors': errors[:100],
        'rows_errors': rows_errors,
        'pii_auto_created': pii_pre.get('pii_auto_created', 0),
        'pii_auto_skipped': pii_pre.get('pii_auto_skipped', 0),
    }


def import_main_mcq_response(df, track_number, progress_callback=None, score_from_sheet=False):
    """
    Import Main MCQ (MCQ Verification) responses from a DataFrame into main_mcq_response.
    Upsert by (track_number, email): update if exists, else insert.
    Cohort 1: score via main_mcq_answer_key. Cohort 2/3: prefer Score column
    (score_from_sheet=True) because question banks differ from Cohort 1.
    Returns dict: total_rows, created, updated, skipped, errors.
    """
    from server.models import db
    from server.utils.main_mcq_answer_key import score_submission as main_score_submission

    IM = _import_models()
    MainMcqResponse = IM.MainMcqResponse

    columns = list(df.columns)
    email_col = _find_leader_email_column(columns)
    if not email_col:
        raise Exception(
            "Could not find a 'Leader Email' column in the Main MCQ sheet. "
            "Please ensure the sheet has a column named 'Leader Email'."
        )
    question_cols = _find_mcq_question_columns(columns, email_col)
    if not question_cols and score_from_sheet:
        question_cols = _find_question_columns_score_to_created(columns)
    if not question_cols:
        raise Exception("Could not find 10 question columns in the Main MCQ sheet.")

    score_col = _find_mcq_column(columns, 'score')
    leader_name_col = _find_mcq_column(columns, 'leader name', 'Leader Name')
    leader_phone_col = _find_mcq_column(columns, 'leader phone', 'Leader Phone')
    team_size_col = _find_mcq_column(columns, 'team size', 'Team size', 'team_size')
    problem_col = _find_mcq_column(columns, 'problem statements', 'Problem Statements', 'problem statement')

    total_rows = len(df)
    created = 0
    updated = 0
    skipped = 0
    errors = []
    rows_errors = []

    from sqlalchemy import func

    pii_pre = _ensure_user_pii_for_leaders(
        df, email_col, name_col=leader_name_col, phone_col=leader_phone_col,
    )

    emails_in_sheet = set()
    try:
        for v in df[email_col].dropna():
            e = str(v).strip().lower()
            if e and '@' in e:
                emails_in_sheet.add(e)
    except Exception:
        pass

    existing = {}
    try:
        for chunk in _chunk_list(emails_in_sheet, IMPORT_EMAIL_IN_CHUNK):
            if not chunk:
                continue
            for mrow in MainMcqResponse.query.filter(
                MainMcqResponse.track_number == track_number,
                func.lower(MainMcqResponse.email).in_(chunk),
            ).all():
                existing[mrow.email.lower()] = mrow
    except Exception:
        pass

    for index, row in df.iterrows():
        email_val = row.get(email_col)
        if pd.isna(email_val):
            skipped += 1
            _record_row_error(errors, rows_errors, _classify_row_error(
                index, None, raw_email='', reason_code='missing_email',
                reason_message='Missing email',
            ))
            continue
        email = str(email_val).strip().lower()
        if not email or '@' not in email:
            skipped += 1
            _record_row_error(errors, rows_errors, _classify_row_error(
                index, None, raw_email=email, reason_code='invalid_email',
                reason_message='Invalid email',
            ))
            continue

        try:
            with db.session.begin_nested():
                def _get(col, default=None):
                    if col is None:
                        return default
                    v = row.get(col)
                    if pd.isna(v):
                        return default
                    return str(v).strip() or default

                data = {'track_number': track_number, 'email': email}
                if leader_name_col:
                    data['leader_name'] = _get(leader_name_col)
                if leader_phone_col:
                    data['leader_phone'] = _get(leader_phone_col)
                if team_size_col:
                    try:
                        v = row.get(team_size_col)
                        data['team_size'] = int(v) if v is not None and not pd.isna(v) else None
                    except (ValueError, TypeError):
                        data['team_size'] = None
                if problem_col:
                    data['problem_statement'] = _get(problem_col)

                for xlsx_col, model_field in question_cols:
                    val = row.get(xlsx_col)
                    if pd.isna(val):
                        data[model_field] = None
                    else:
                        data[model_field] = str(val).strip() or None

                if score_from_sheet:
                    num, den = _parse_mcq_score_slash(row.get(score_col)) if score_col else (None, None)
                    if num is not None and den and den > 0:
                        data['score'] = min(10, max(0, int(round(num * 10.0 / float(den)))))
                    elif num is not None:
                        data['score'] = min(10, max(0, num))
                    else:
                        data['score'] = None
                else:
                    auto = main_score_submission(
                        track_number,
                        data.get('question_1'), data.get('question_2'), data.get('question_3'), data.get('question_4'),
                        data.get('question_5'), data.get('question_6'), data.get('question_7'), data.get('question_8'),
                        data.get('question_9'), data.get('question_10'),
                    )
                    data['score'] = auto['correct_count']

                rec = existing.get(email)
                if rec:
                    for k, v in data.items():
                        if k == 'email':
                            continue
                        setattr(rec, k, v)
                    updated += 1
                else:
                    rec = MainMcqResponse(**data)
                    db.session.add(rec)
                    existing[email] = rec
                    created += 1

            if progress_callback:
                try:
                    progress_callback(created, updated, skipped)
                except Exception:
                    pass

        except Exception as e:
            skipped += 1
            _record_row_error(errors, rows_errors, _classify_row_error(
                index, e, raw_email=email,
            ))

    try:
        db.session.commit()
    except Exception:
        db.session.rollback()

    return {
        'total_rows': total_rows,
        'created': created,
        'updated': updated,
        'skipped': skipped,
        'errors': errors[:100],
        'rows_errors': rows_errors,
        'pii_auto_created': pii_pre.get('pii_auto_created', 0),
        'pii_auto_skipped': pii_pre.get('pii_auto_skipped', 0),
    }
